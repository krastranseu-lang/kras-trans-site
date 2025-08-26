from openpyxl import Workbook, load_workbook
import sys
sys.path.append('tools')
import build


def test_markdown_heading_compilation(tmp_path):
    # Create temporary XLSX with Markdown heading
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pages'
    ws.append(['body_md'])
    ws.append(['## XXX'])
    xlsx_path = tmp_path / 'sample.xlsx'
    wb.save(xlsx_path)

    # Load rows using builder helper
    sheet = load_workbook(xlsx_path).active
    rows = build.load_rows(sheet)
    assert rows[0]['body_md'] == '## XXX'

    # Compile Markdown to HTML
    html = build.md_to_html(rows[0]['body_md'])
    assert '<h2>XXX</h2>' in html
    assert '##' not in html

