from pathlib import Path
import subprocess
import sys
from openpyxl import load_workbook


def test_blocks_filtering(tmp_path):
    src = Path('data/cms/menu.xlsx')
    backup = tmp_path / 'menu.xlsx.bak'
    backup.write_bytes(src.read_bytes())
    try:
        wb = load_workbook(src)
        ws = wb['Blocks']
        header = [cell.value for cell in ws[1]]
        header_lc = [str(h or '').strip().lower() for h in header]

        def idx(name: str):
            try:
                return header_lc.index(name)
            except ValueError:
                return None

        row = [''] * len(header)
        i = idx('block')
        if i is not None:
            row[i] = 'testblock'
        i = idx('lang')
        if i is not None:
            row[i] = 'PL '
        i = idx('page')
        if i is not None:
            row[i] = ' home '
        i = idx('body_md')
        if i is not None:
            row[i] = 'Testing block body'
        i = idx('enabled')
        if i is not None:
            row[i] = '1'
        ws.append(row)
        wb.save(src)

        subprocess.run([sys.executable, 'tools/build.py'], check=True)
        html = Path('dist/pl/index.html').read_text('utf-8')
        assert 'Testing block body' in html
    finally:
        src.write_bytes(backup.read_bytes())
        subprocess.run([sys.executable, 'tools/build.py'], check=True)
