"""Content synchronization tests."""

import json
from pathlib import Path
from urllib.parse import urlsplit

import openpyxl
import pytest
from bs4 import BeautifulSoup
import sys
sys.path.append('tools')
import cms_ingest

XLSX_PATH = Path('data/cms/menu.xlsx')
DIST = Path('dist')


def truthy(val):
    s = str(val).strip().lower()
    return s in {'1', 'true', 'yes', 'tak', 'on', 'prawda'}


def cell_str(val):
    """Normalize XLSX cell value to string."""
    if val is None:
        return ''
    return str(val).strip()


FIELD_EXTRACTORS = {
    'title': lambda soup: soup.title.string.strip() if soup.title and soup.title.string else '',
    'h1': lambda soup: soup.find('h1').get_text(strip=True) if soup.find('h1') else '',
    'lead': lambda soup: (soup.select_one('#hero-lead').get_text(strip=True)
                          if soup.select_one('#hero-lead') else ''),
    'cta_label': lambda soup: (soup.select_one('.cta-row a.btn-primary').get_text(strip=True)
                               if soup.select_one('.cta-row a.btn-primary') else ''),
    'cta_secondary': lambda soup: (soup.select_one('.cta-row a.btn-ghost[href^="/"]').get_text(strip=True)
                                   if soup.select_one('.cta-row a.btn-ghost[href^="/"]') else ''),
    'cta_phone': lambda soup: (soup.select_one('.cta-row a[href^="tel:"]').get_text(strip=True)
                               if soup.select_one('.cta-row a[href^="tel:"]') else ''),
    'meta_desc': lambda soup: (soup.select_one('meta[name="description"]')['content'].strip()
                               if soup.select_one('meta[name="description"]') else ''),
    'canonical_path': lambda soup: (urlsplit(soup.find('link', rel='canonical')['href']).path
                                    if soup.find('link', rel='canonical') and soup.find('link', rel='canonical').get('href')
                                    else ''),
}


def load_sheet(name):
    if not XLSX_PATH.exists():
        pytest.skip("CMS sheet missing")
    try:
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    except Exception:
        pytest.skip("CMS sheet unavailable")
    return wb[name]


def test_pages_content_matches_cms():
    ws = load_sheet('Pages')
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h or '').strip().lower() for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    for r in rows[1:]:
        if not r:
            continue
        lang = cell_str(r[idx['lang']]).lower()
        slug = cell_str(r[idx['slug']]).strip('/')
        publish = truthy(r[idx['publish']]) if 'publish' in idx else True
        if not publish:
            continue
        path = DIST / lang
        if slug:
            path /= slug
        path /= 'index.html'
        if not path.exists():
            continue
        soup = BeautifulSoup(path.read_text(encoding='utf-8'), 'lxml')
        expected = {
            'title': cell_str(r[idx.get('seo_title')]) or cell_str(r[idx.get('title')]),
            'h1': cell_str(r[idx.get('h1')]),
            'lead': cell_str(r[idx.get('lead')]),
            'cta_label': cell_str(r[idx.get('cta_label')]),
            'cta_secondary': cell_str(r[idx.get('cta_secondary')]),
            'cta_phone': cell_str(r[idx.get('cta_phone')]),
            'meta_desc': cell_str(r[idx.get('meta_desc')]),
            'canonical_path': cell_str(r[idx.get('canonical_path')]),
        }
        actual = {k: extractor(soup) for k, extractor in FIELD_EXTRACTORS.items()}
        for field, exp in expected.items():
            act = actual.get(field, '')
            if exp and act:
                if field == 'title':
                    _, _, tail = exp.partition('|')
                    assert act.endswith(tail.strip()), (
                        f"title mismatch for {path}: {act!r} != {exp!r}"
                    )
                elif field == 'h1':
                    words = [w.lower() for w in exp.split() if len(w) > 2]
                    assert any(w in act.lower() for w in words), (
                        f"h1 mismatch for {path}: {act!r} != {exp!r}"
                    )
                elif field == 'meta_desc':
                    words = [w.lower() for w in exp.split() if len(w) > 2]
                    assert any(w in act.lower() for w in words), (
                        f"meta_desc mismatch for {path}: {act!r} != {exp!r}"
                    )
                else:
                    assert act == exp, f"{field} mismatch for {path}: {act!r} != {exp!r}"


def test_home_has_blocks():
    ws = load_sheet('Blocks')
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        pytest.skip('no block rows')
    header = [str(h or '').strip().lower() for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    if 'lang' not in idx or 'page' not in idx:
        pytest.skip('unsupported blocks sheet')
    target_lang = 'pl'
    found = False
    for r in rows[1:]:
        if not r:
            continue
        lang = cell_str(r[idx['lang']]).lower()
        page = cell_str(r[idx['page']]).lower()
        enabled = truthy(r[idx.get('enabled')]) if 'enabled' in idx else True
        if lang == target_lang and page == 'home' and enabled:
            found = True
            break
    if not found:
        pytest.skip('no enabled home blocks for lang')

    import openpyxl as ox
    wb = ox.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    cms = cms_ingest.load_all(Path('data')/ 'cms')
    routes = cms.get('routes', {})
    blocks = cms_ingest.load_blocks_for_lang(wb, target_lang, routes)
    assert blocks and len(blocks) > 0


def test_nav_menu_matches_cms():
    ws = load_sheet('Nav')
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h or '').strip().lower() for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    nav_expected = {}
    for r in rows[1:]:
        if not r:
            continue
        lang = (r[idx['lang']] or '').strip().lower()
        enabled = truthy(r[idx.get('enabled', 0)]) if 'enabled' in idx else True
        parent = (r[idx.get('parent')] or '').strip()
        label = (r[idx['label']] or '').strip()
        href = (r[idx['href']] or '').strip()
        order = r[idx.get('order')] or 0
        if not enabled or parent:
            continue
        nav_expected.setdefault(lang, []).append((order, label, href))
    for lang, items in nav_expected.items():
        bundle_path = DIST / 'assets' / 'data' / 'menu' / f'bundle_{lang}.json'
        assert bundle_path.exists(), f'missing bundle {bundle_path}'
        data = json.loads(bundle_path.read_text(encoding='utf-8'))
        actual_items = [
            (it.get('order', 0), it['label'], it['href'])
            for it in data.get('items', [])
        ]
        assert sorted(actual_items) == sorted(items), f"nav mismatch for {lang}"
