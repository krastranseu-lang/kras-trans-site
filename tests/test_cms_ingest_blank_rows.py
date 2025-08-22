import sys
from pathlib import Path

import openpyxl

# Ensure project root on Python path for imports
sys.path.append(str(Path(__file__).resolve().parents[1]))

from tools.cms_ingest import load_all


def test_pages_ignore_trailing_blanks(tmp_path, monkeypatch):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "pages"
    ws.append(["lang", "publish", "slug", "template"])
    ws.append(["pl", "1", "/foo", "page.html"])
    ws.append(["pl", "", "/bar", "page.html"])
    for _ in range(1000):
        ws.append([None, None, None, None])
    src = tmp_path / "test.xlsx"
    wb.save(src)

    count = 0
    original_iter_rows = openpyxl.worksheet.worksheet.Worksheet.iter_rows

    def counting_iter_rows(self, *args, **kwargs):
        nonlocal count
        for row in original_iter_rows(self, *args, **kwargs):
            count += 1
            yield row

    monkeypatch.setattr(openpyxl.worksheet.worksheet.Worksheet, "iter_rows", counting_iter_rows)

    result = load_all(cms_root=tmp_path, explicit_src=src)
    assert [p["slugKey"] for p in result["pages_rows"]] == ["foo"]
    assert count < 50

