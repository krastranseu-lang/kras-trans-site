import sys
from pathlib import Path

import openpyxl

# Ensure project root is on the Python path for imports
sys.path.append(str(Path(__file__).resolve().parents[1]))

from tools.cms_ingest import load_all


def test_slug_without_leading_slash(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "pages"
    ws.append(["lang", "publish", "slug", "template"])
    ws.append(["ua", "1", "ua/tsiny", "page.html"])
    src = tmp_path / "test.xlsx"
    wb.save(src)

    result = load_all(cms_root=tmp_path, explicit_src=src)
    assert result["pages_rows"][0]["slug"] == "tsiny"

def test_missing_required_fields_warn(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "pages"
    ws.append(["lang", "publish", "slug", "template", "h1", "title", "body_md"])
    ws.append(["pl", "1", "/foo", "page.html", "", "", ""])

    src = tmp_path / "test.xlsx"
    wb.save(src)

    result = load_all(cms_root=tmp_path, explicit_src=src)
    warnings = result.get("warnings") or []
    slug = result["pages_rows"][0]["slugKey"]
    assert f"[cms_ingest] page '{slug}' missing h1" in warnings
    assert f"[cms_ingest] page '{slug}' missing title" in warnings
    assert f"[cms_ingest] page '{slug}' missing body_md" in warnings


def test_rows_respect_calculated_dimension():
    from tools.cms_ingest import _rows

    class FakeWs:
        title = "Fake"
        max_row = 9999

        def calculate_dimension(self):
            return "A1:B2"  # only two rows (header + one data row)

        def iter_rows(self, values_only=True, max_row=None):
            assert max_row == 2
            data = [
                ("h1", "h2"),
                ("r1c1", "r1c2"),
                ("extra", "row"),
            ]
            for row in data[:max_row]:
                yield row

    rows = list(_rows(FakeWs()))
    assert len(rows) == 1
    assert rows[0][1] == ["r1c1", "r1c2"]
