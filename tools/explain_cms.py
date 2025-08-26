from pathlib import Path
from openpyxl import load_workbook
import sys

SRC = Path('data/cms/menu.xlsx')
OUT = Path('docs/CMS-CONTRACT.md')

REQ = {
  'Pages': ['lang','type','slug','slugKey','template','publish','order','h1','title','seo_title','meta_desc','lead','cta_label','body_md'],
  'Routes': ['slugKey','pl','en','de','fr','it','ru','ua'],
  'Nav': ['lang','label','href','parent','order','col','enabled','nav.logo.src','nav.logo.alt','nav.cta.label','nav.cta.slugKey','nav.social.ig','nav.social.li','nav.social.fb'],
  'Blocks': ['block','lang','page','type','title','desc','href','order','enabled'],
}

def main():
  wb = load_workbook(SRC, data_only=True)
  lines = []
  lines.append(f'# CMS CONTRACT — {SRC}\n')
  for sheet, cols in REQ.items():
    ws = wb[sheet]
    header = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(max_row=1))]
    miss = [c for c in cols if c not in header]
    lines.append(f'## {sheet}\nWymagane kolumny:\n- ' + '\n- '.join(cols))
    if miss:
      lines.append(f'\n❌ Brakujące kolumny: {", ".join(miss)}\n')
    else:
      lines.append('\n✅ Kolumny OK\n')
  OUT.parent.mkdir(parents=True, exist_ok=True)
  OUT.write_text('\n'.join(lines), encoding='utf-8')
  print(f'[ok] Wygenerowano {OUT}')

if __name__ == '__main__':
  main()
