#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Menu Builder for Kras-Trans.
- Reads data/cms/cms.json OR cms.csv OR cms.xlsx (first sheet).
- Expected columns: lang, label, href, parent, order, col, enabled
- Produces per-language bundles and pre-rendered HTML for instant SSR.
- Labels must be unique within each language; duplicates emit a warning and are ignored.
"""
from __future__ import annotations
import csv, json, hashlib, re, unicodedata, datetime
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional

SAFE_PROTOCOLS = ("http://", "https://", "/", "#", "mailto:", "tel:")

def _slugify(s: str) -> str:
    s = unicodedata.normalize('NFKD', s).encode('ascii','ignore').decode('ascii')
    s = re.sub(r'[^a-zA-Z0-9]+', '-', s).strip('-').lower()
    return s or "item"

def _to_bool(v: Any) -> bool:
    if isinstance(v, bool): return v
    if v is None: return False
    s = str(v).strip().lower()
    return s in {"1","true","t","yes","y","on","prawda","tak"}

def _to_int(v: Any, default: int) -> int:
    try:
        return int(float(str(v).strip()))
    except Exception:
        return default

def _sanitize_label(s: Any) -> str:
    return (str(s or "").strip())[:120]

def _sanitize_href(href: Any, lang: str, label: str) -> str:
    h = str(href or "").strip()
    if not h:
        return f"/{lang}/{_slugify(label)}/"
    # allowlist protocols
    ok = h.startswith(SAFE_PROTOCOLS)
    return h if ok else f"/{lang}/{_slugify(label)}/"

def _load_json(p: Path) -> List[Dict[str, Any]]:
    return json.loads(p.read_text(encoding="utf-8"))

def _load_csv(p: Path) -> List[Dict[str, Any]]:
    rows = []
    with p.open("r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows.append(row)
    return rows

def _load_xlsx(p: Path) -> List[Dict[str, Any]]:
    try:
        import openpyxl  # optional
    except ImportError as e:
        raise SystemExit("Do odczytu XLSX zainstaluj: pip install openpyxl  — lub zapisz CMS jako JSON/CSV.") from e
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h:i for i,h in enumerate(headers)}
    req = ["lang","label","href","parent","order","col","enabled"]
    for r in req:
        if r not in idx:
            raise SystemExit(f"Brak kolumny '{r}' w {p.name}")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {h: row[idx[h]] for h in req}
        rows.append(d)
    return rows

def load_cms(cms_dir: Path) -> List[Dict[str, Any]]:
    """
    Ładuje menu z arkusza/JSON/CSV. Szuka w:
    - data/cms/  (standard)
    - data/      (root, gdy ktoś trzyma pliki bez podfolderu)
    Obsługuje nazwy: cms.json, CMS.json, cms.csv, CMS.csv, cms.xlsx, CMS.xlsx.
    Priorytet: XLSX > JSON > CSV (jeśli znajdzie kilka — bierze pierwszy wg kolejności poniżej).
    """
    cms_dir = Path(cms_dir)
    roots = [cms_dir]
    if cms_dir.parent:
        roots.append(cms_dir.parent)  # np. data/

    # Kolejność preferencji: XLSX → JSON → CSV
    names = [
        "CMS.xlsx", "cms.xlsx",
        "cms.json", "CMS.json",
        "cms.csv",  "CMS.csv",
    ]

    src: Optional[Path] = None
    for root in roots:
        for name in names:
            p = root / name
            if p.exists():
                src = p
                break
        if src:
            break

    if not src:
        # brak plików – zwróć pustą listę, builder użyje fallbacku
        return []

    # Wczytaj wg rozszerzenia
    ext = src.suffix.lower()
    print(f"[menu_builder] Using CMS source: {src}")
    if ext == ".json":
        raw = _load_json(src)
    elif ext == ".csv":
        raw = _load_csv(src)
    elif ext == ".xlsx":
        raw = _load_xlsx(src)
    else:
        return []

    # Normalizacja jak dotąd + kontrola duplikatów etykiet per język
    norm: List[Dict[str, Any]] = []
    labels_by_lang: Dict[str, set] = {}
    for r in raw:
        lang = str(r.get("lang", "pl")).strip().lower()
        label = _sanitize_label(r.get("label"))
        parent = _sanitize_label(r.get("parent"))
        href = _sanitize_href(r.get("href"), lang, label)
        order = _to_int(r.get("order"), 999)
        col = _to_int(r.get("col"), 1)
        enabled = _to_bool(r.get("enabled"))
        if not label:
            continue

        seen = labels_by_lang.setdefault(lang, set())
        if label in seen:
            print(f"[menu_builder] WARNING: duplicate label '{label}' for lang '{lang}' — skipping")
            continue
        seen.add(label)

        norm.append({
            "lang": lang,
            "label": label,
            "href": href,
            "parent": parent,
            "order": order,
            "col": col,
            "enabled": enabled,
        })

    norm = [r for r in norm if r["enabled"]]
    return norm

def build_bundle_for_lang(rows: List[Dict[str, Any]], lang: str) -> Dict[str, Any]:
    LR = [r for r in rows if r["lang"] == lang]
    if not LR:
        return {"lang": lang, "version": "sha256:0"*8, "generated_at": datetime.datetime.utcnow().isoformat()+"Z", "items": []}

    # Partition to mains and children
    mains = [r for r in LR if not r["parent"]]
    mains.sort(key=lambda r: (r["order"], r["label"].lower()))
    by_label = {r["label"]: r for r in LR}
    children = [r for r in LR if r["parent"]]

    # Group children under parent
    children_by_parent: Dict[str, List[Dict[str, Any]]] = {}
    for c in children:
        if c["parent"] in by_label:
            children_by_parent.setdefault(c["parent"], []).append(c)

    items = []
    for m in mains:
        node = {"label": m["label"], "href": m["href"], "order": m["order"]}
        kids = children_by_parent.get(m["label"], [])
        if kids:
            kids.sort(key=lambda r: (r["order"], r["label"].lower()))
            # Columns
            cols: Dict[int, List[Dict[str, Any]]] = {}
            for ch in kids:
                cols.setdefault(ch["col"], []).append({"label": ch["label"], "href": ch["href"], "order": ch["order"]})
            # ensure sort inside each col
            for k in cols:
                cols[k].sort(key=lambda r: (r["order"], r["label"].lower()))
            ordered_cols = [cols[k] for k in sorted(cols)]
            node["cols"] = ordered_cols
        items.append(node)

    payload = {
        "lang": lang,
        "generated_at": datetime.datetime.utcnow().isoformat()+"Z",
        "items": items
    }
    # Version: hash of items only (stable)
    h = hashlib.sha256(json.dumps(items, sort_keys=True, ensure_ascii=False, separators=(",",":")).encode("utf-8")).hexdigest()
    payload["version"] = f"sha256:{h}"
    return payload

def escape_html(s: str) -> str:
    return (s.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
             .replace('"',"&quot;").replace("'","&#39;"))

def render_nav_html(bundle: Dict[str, Any]) -> str:
    """Render <li> items for <ul class='nav-list'>"""
    li = []
    for it in sorted(bundle.get("items", []), key=lambda i: (i.get("order",999), i.get("label","").lower())):
        label = escape_html(it["label"])
        href = escape_html(it.get("href","/"))
        cols = it.get("cols")
        if cols:
            mega_id = f"mega-{_slugify(label)}"
            parts = []
            parts.append(f'<li class="has-mega">')
            parts.append(f'  <button class="mega-toggle" aria-expanded="false" aria-controls="{mega_id}">{label}</button>')
            parts.append(f'  <div id="{mega_id}" class="mega" role="dialog" aria-label="{label}" aria-modal="false">')
            parts.append(f'    <div class="mega-grid">')
            for col in cols:
                parts.append('      <div class="mega-col"><ul>')
                for ch in col:
                    ch_label = escape_html(ch["label"])
                    ch_href = escape_html(ch["href"])
                    parts.append(f'        <li><a href="{ch_href}">{ch_label}</a></li>')
                parts.append('      </ul></div>')
            parts.append(f'    </div>')
            parts.append(f'  </div>')
            parts.append(f'</li>')
            li.append("\n".join(parts))
        else:
            li.append(f'<li><a href="{href}">{label}</a></li>')
    return "\n".join(li)

def build_all(cms_dir: Path, languages: List[str]) -> Tuple[Dict[str, Any], Dict[str, str]]:
    """Return (bundles_by_lang, html_by_lang)"""
    rows = load_cms(cms_dir)
    bundles: Dict[str, Any] = {}
    html_by_lang: Dict[str, str] = {}
    for lang in languages:
        b = build_bundle_for_lang(rows, lang)
        bundles[lang] = b
        html_by_lang[lang] = render_nav_html(b)
    return bundles, html_by_lang
