#!/usr/bin/env python3
"""Generate snapshot report for CMS XLSX file."""
import json
import os
import sys
from collections import Counter
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

LANG_HEADERS = {"lang", "język", "jezyk"}
TRUE_VALUES = {"true", "1", "yes", "tak", "on"}

def _truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value == 1
    if isinstance(value, str):
        return value.strip().lower() in TRUE_VALUES
    return False

def main(path: Optional[str] = None) -> None:
    # Resolve path from argument or environment variable
    if path is None:
        path = os.environ.get("CMS_SOURCE")
    if not path:
        print("Usage: python tools/cms_snapshot.py <xlsx_path>", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(path, read_only=True, data_only=True)

    sheets: List[Dict[str, Any]] = []
    rows_per_lang: Counter[str] = Counter()
    published_per_lang: Counter[str] = Counter()

    for ws in wb.worksheets:
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        sheets.append({"name": ws.title, "headers": headers})
        print(f"{ws.title}: {headers}")

        header_map = {
            (str(h).casefold() if isinstance(h, str) else None): idx
            for idx, h in enumerate(headers)
        }

        # locate language and publish columns
        lang_idx = None
        for alias in LANG_HEADERS:
            idx = header_map.get(alias.casefold())
            if idx is not None:
                lang_idx = idx
                break
        publish_idx = header_map.get("publish")

        if lang_idx is None:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if lang_idx >= len(row):
                continue
            lang_val = row[lang_idx]
            if not lang_val:
                continue
            lang = str(lang_val).strip()
            if not lang.isalpha():
                continue
            rows_per_lang[lang] += 1
            if publish_idx is not None and publish_idx < len(row):
                if _truthy(row[publish_idx]):
                    published_per_lang[lang] += 1

    report = {
        "sheets": sheets,
        "rows_per_lang": dict(rows_per_lang),
        "published_per_lang": dict(published_per_lang),
    }

    with open("sheet_report.json", "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    arg = sys.argv[1] if len(sys.argv) > 1 else None
    main(arg)
