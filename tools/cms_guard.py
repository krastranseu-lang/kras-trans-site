#!/usr/bin/env python3
from __future__ import annotations
from pathlib import Path
from typing import List
import sys


def validate(schema_path: Path, xlsx_path: Path) -> None:
    import openpyxl, yaml
    _ = yaml.safe_load(schema_path.read_text(encoding="utf-8")) if schema_path.exists() else None
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    def norm(s: str) -> str:
        return (s or "").strip().lower()

    content_like = 0
    recognized = 0

    for ws in wb.worksheets:
        start_cell, end_cell = ws.calculate_dimension().split(":")
        max_row = int("".join(filter(str.isdigit, end_cell)))
        hdr = [str(c or "").strip() for c in next(ws.iter_rows(values_only=True, max_row=max_row))]
        hdr_lc = [norm(h) for h in hdr]
        print(f"[sheet] {ws.title}: {hdr}")

        is_pages = "lang" in hdr_lc and "publish" in hdr_lc and ("slug" in hdr_lc or "slugkey" in hdr_lc) and "template" in hdr_lc
        is_menu = "lang" in hdr_lc and "publish" in hdr_lc and "label" in hdr_lc and "href" in hdr_lc
        is_meta = "lang" in hdr_lc and "key" in hdr_lc
        is_blocks = "lang" in hdr_lc and ("html" in hdr_lc or "body" in hdr_lc)

        if "lang" in hdr_lc and "publish" in hdr_lc:
            if any(c in hdr_lc for c in ["slug", "slugkey", "template", "label", "href", "enabled", "html", "body", "key"]):
                content_like += 1

        if is_pages or is_menu or is_meta or is_blocks:
            recognized += 1

    print(f"[summary] content-like={content_like}, recognized={recognized}")
    if recognized < content_like:
        print("[cms_guard] ERROR: some content-like sheets not recognized")
        sys.exit(1)

    # zapis debug
    try:
        import json, os
        def _hdr(ws):
            start_cell, end_cell = ws.calculate_dimension().split(":")
            max_row = int("".join(filter(str.isdigit, end_cell)))
            return [str(c or "") for c in next(ws.iter_rows(values_only=True, max_row=max_row))]
        rep = {"sheets":[{"name":ws.title,"headers":_hdr(ws)} for ws in wb.worksheets]}
        Path("sheet_report.json").write_text(json.dumps(rep, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def main() -> int:
    if len(sys.argv) < 3:
        print("Usage: cms_guard.py <schema.yml> <xlsx>")
        return 1
    schema = Path(sys.argv[1])
    xlsx = Path(sys.argv[2])
    validate(schema, xlsx)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
