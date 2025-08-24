# tools/cms_ingest.py
from __future__ import annotations
from pathlib import Path
from typing import Dict, Any, List, Optional
import re

LANG_RE = re.compile(r"^/([a-z]{2})(?:/([^?#]*))?/?$")


def split_slug(slug: str):
    """
    Zwraca (lang, rel) dla slugów w formacie '/pl/kontakt/'.
    Jeśli slug pusty → (None, '').
    """
    s = (slug or "").strip()
    m = LANG_RE.match(s)
    if not m:
        return None, ""
    lang = m.group(1)
    rel = (m.group(2) or "").strip("/")
    return lang, rel

SYN = {
  # kolumny dla arkuszy z definicjami stron
  "pages": {
    "lang": ["lang", "język", "jezyk"],
    "publish": ["publish", "enabled", "widoczne", "visible", "aktywny", "active"],
    "slug": ["slug", "url", "path"],
    "key": ["slugkey", "slug_key", "key", "page", "route"],
    "parent": ["parent", "parent_key", "parentslug", "parent_slug", "parentkey"],
    "template": ["template", "tpl", "szablon"],
    "order": ["order", "kolej", "sort", "kolejność", "kolejnosc", "poz"],
    "title": ["title", "meta_title", "tytuł", "tytul"],
    "seo_title": ["seo_title", "title", "meta_title"],
    "description": ["description", "meta_desc", "opis", "desc"],
    "og_image": ["og_image", "og", "image", "obraz", "grafika"],
    "canonical": ["canonical", "kanoniczny", "canonical_url"],
  },
  # menu nawigacyjne
  "menu": {
    "lang":["lang","język","jezyk"],
    "label":["label","nazwa","etykieta","tekst"],
    "href":["href","url","link"],
    "parent":["parent","rodzic","grupa"],
    "order":["order","kolej","sort","kolejność","kolejnosc","poz"],
    "col":["col","kol","kolumna","column"],
    "enabled":["enabled","visible","widoczne","on","aktywny","active"]
  },
  "meta": {
    "lang":["lang","język","jezyk"],
    "key":["key","slugkey","page","slug","strona","zakladka","route"],
    "title":["title","meta_title","tytuł","tytul"],
    "description":["description","meta_desc","opis","desc"],
    "og_image":["og_image","og","image","obraz","grafika"]
  },
  "blocks": {
    "lang":["lang","język","jezyk"],
    "key":["key","slugkey","page","slug","strona","zakladka","route"],
    "section":["section","sekcja","blok","area","part"],
    "path":["path","ścieżka","sciezka"],
    "html":["html","content_html"],
    "title":["title","naglowek","header","h1","h2","h3"],
    "body":["body","tekst","content","markdown","md","body_md","body_html"],
    "cta_label":["cta_label","cta","button","przycisk","label"],
    "cta_href":["cta_href","cta_link","button_link","href","link"]
  },
  "faq": {
    "lang":["lang","język","jezyk"],
    "q":["q","question","pytanie"],
    "a":["a","answer","odp","odpowiedz","odpowiedź"],
    "page_slug":["page_slug","slugkey","slug","page","strona"],
    "order":["order","kolej","sort"],
    "enabled":["enabled","visible","widoczne","on","active","aktywny"]
  },
  "props": {
    "key":["key","prop","nazwa"],
    "lang":["lang","język","jezyk"],
    "value":["value","wartosc","wartość","val","content"]
  },
  "blog": {
    "lang": ["lang", "język", "jezyk"],
    "publish": ["publish", "enabled", "widoczne", "visible", "aktywny", "active"],
    "slug": ["slug", "url", "path"],
    "title": ["title", "tytuł", "tytul"],
    "h1": ["h1", "header", "naglowek"],
    "lead": ["lead", "intro", "opis", "description", "meta_desc"],
    "body": ["body", "html", "content", "markdown", "md"],
    "hero_image": ["hero_image", "image", "img"],
    "published_at": ["published_at", "date", "published", "data"],
    "tags": ["tags", "tagi"],
    "categories": ["categories", "kategorie"]
  },
  "routes": {
    "slugkey": ["slugkey", "slug_key", "key", "route", "slug"],
  },
  "strings": {
    "key": ["key", "nazwa"],
  },
  "media": {
    "src": ["src", "path", "url"],
    "alt": ["alt", "opis", "description"],
    "title": ["title", "tytuł", "tytul"],
  },
  "company": {
    "name": ["name", "nazwa"],
    "street_address": ["street_address", "address", "street", "adres"],
    "postal_code": ["postal_code", "post_code", "zip", "kod"],
    "city": ["city", "miasto"],
    "telephone": ["telephone", "phone", "tel"],
    "email": ["email", "mail"],
    "same_as": ["same_as", "social", "socials", "links"],
  },
  "redirects": {
    "from": ["from", "src", "source"],
    "to": ["to", "dst", "dest", "target"],
  }
}

LOCALES = {"pl", "en", "de", "fr", "it", "ru", "ua"}

def _lower(s:str) -> str: return (s or "").strip().lower()

def _norm_slug(lang: str, raw: str) -> str:
    s = (raw or "").strip()
    cut = len(s)
    for ch in ("#", "?"):
        if ch in s:
            cut = min(cut, s.index(ch))
    s = s[:cut]
    if s.startswith(f"/{lang}/"):
        s = s[len(f"/{lang}/"):]
    s = "/".join([p.strip() for p in s.split("/") if p.strip()])
    return s + ("" if s.endswith("/") or s == "" else "/")
def _map_headers(headers: List[str], syn: Dict[str,List[str]]) -> Dict[str,int]:
    hl=[_lower(h) for h in headers]; out={}
    for want, aliases in syn.items():
        for a in aliases:
            if _lower(a) in hl: out[want]=hl.index(_lower(a)); break
    return out

def _read_xlsx(path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return {ws.title: ws for ws in wb.worksheets}

def _log_path(path: Path) -> Path:
    """Ensure ``path`` is :class:`Path` and log it."""
    p = Path(path)
    print(f"[cms_ingest] loading {p}")
    return p

def _rows(ws):
    it = ws.iter_rows(values_only=True)
    headers = [str(x or "").strip() for x in next(it)]
    for row in it:
        yield headers, [("" if v is None else str(v).strip()) for v in row]

def _find_sheet(sheets, group):
    best = None; best_score=-1
    for name, ws in sheets.items():
        try:
            it = ws.iter_rows(values_only=True)
            headers = [str(x or "").strip() for x in next(it)]
            m = _map_headers(headers, SYN[group])
            score = len(m)
            if score > best_score:
                best = name; best_score = score
        except Exception: 
            pass
    return best if best_score >= 3 else None

def load_all(cms_root: Path, explicit_src: Optional[Path] = None) -> Dict[str, Any]:
    """Wczytuje wszystkie arkusze XLSX i klasyfikuje je podobnie jak ``cms_guard``.

    Wynik łączy dane z wielu arkuszy (union) i zwraca słownik zawierający
    klucze wymagane przez zadanie.
    """

    cms_root = _log_path(cms_root)
    report: List[str] = []

    # wybór źródła
    src = (
        explicit_src
        if explicit_src and explicit_src.exists()
        else (cms_root / "menu.xlsx" if (cms_root / "menu.xlsx").exists() else None)
    )
    if not src:
        return {
            "pages_rows": [],
            "page_routes": {},
            "routes": {},
            "menu_rows": [],
            "page_meta": {},
            "blocks": {},
            "blog_rows": [],
            "strings": [],
            "media": [],
            "company": [],
            "redirects": [],
            "collections": {},
            "report": "[cms] no source",
        }

    report.append(f"[cms_ingest] source: {src}")

    import openpyxl

    try:
        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    except Exception as e:
        report.append(f"[cms_ingest] warn: {e}")
        return {
            "pages_rows": [],
            "page_routes": {},
            "routes": {},
            "menu_rows": [],
            "page_meta": {},
            "blocks": {},
            "blog_rows": [],
            "strings": [],
            "media": [],
            "company": [],
            "redirects": [],
            "collections": {},
            "report": "\n".join(report),
        }

    def _norm(s):
        return (str(s or "")).strip().lower()

    def _row_empty(row):
        return row is None or all((c is None or str(c).strip() == "") for c in row)

    def _idx(headers_lc, name):
        try:
            return headers_lc.index(name)
        except ValueError:
            return -1

    def _cell(row, headers_lc, name):
        i = _idx(headers_lc, name)
        if i < 0:
            return ""
        v = row[i] if i < len(row) else ""
        return "" if v is None else str(v).strip()

    def truthy(v: str) -> bool:
        return _norm(v) in {"1", "true", "tak", "yes", "on", "prawda"}

    # akumulatory
    menu_rows: List[Dict[str, Any]] = []
    page_meta: Dict[str, Dict[str, Dict[str, Any]]] = {}
    blocks: Dict[str, Dict[str, Dict[str, Any]]] = {}
    pages_rows: List[Dict[str, Any]] = []
    routes: Dict[str, Dict[str, str]] = {}
    collections: Dict[str, Dict[str, List[Dict[str, Any]]]] = {}
    blog_rows: List[Dict[str, Any]] = []
    strings_rows: List[Dict[str, Any]] = []
    media_rows: List[Dict[str, Any]] = []
    company_rows: List[Dict[str, Any]] = []
    redirect_rows: List[Dict[str, Any]] = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        hdr = [str(x or "").strip() for x in rows[0]]
        hdr_lc = [_norm(h) for h in hdr]
        report.append(f"[sheet] {ws.title}: {hdr}")

        m_pages = _map_headers(hdr, SYN.get("pages", {}))
        m_menu = _map_headers(hdr, SYN["menu"])
        m_meta = _map_headers(hdr, SYN["meta"])
        m_blocks = _map_headers(hdr, SYN["blocks"])
        m_blog = _map_headers(hdr, SYN.get("blog", {}))
        m_routes_sheet = _map_headers(hdr, SYN.get("routes", {}))
        m_strings = _map_headers(hdr, SYN.get("strings", {}))
        m_media = _map_headers(hdr, SYN.get("media", {}))
        m_company = _map_headers(hdr, SYN.get("company", {}))
        m_redirects = _map_headers(hdr, SYN.get("redirects", {}))

        data_rows = rows[1:]

        is_pages = (
            "lang" in m_pages
            and "publish" in m_pages
            and "template" in m_pages
            and ("slug" in m_pages or "key" in m_pages)
        )
        is_menu = all(k in m_menu for k in ("lang", "label", "href", "enabled"))
        is_meta = all(k in m_meta for k in ("lang", "key"))
        is_blocks = ("lang" in m_blocks) and ("html" in m_blocks or "body" in m_blocks)
        is_blog = ("lang" in m_blog) and ("slug" in m_blog or "title" in m_blog)
        is_routes_sheet = "slugkey" in m_routes_sheet
        is_strings = "key" in m_strings and any(h in LOCALES for h in hdr_lc)
        is_media = bool(m_media)
        is_company = bool(m_company)
        is_redirects = all(k in m_redirects for k in ("from", "to"))
        klass = (
            "pages"
            if is_pages
            else "menu"
            if is_menu
            else "meta"
            if is_meta
            else "blocks"
            if is_blocks
            else "blog"
            if is_blog
            else "routes"
            if is_routes_sheet
            else "strings"
            if is_strings
            else "media"
            if is_media
            else "company"
            if is_company
            else "redirects"
            if is_redirects
            else "collection"
        )

        if is_pages:
            report.append(f"[detect] pages-like: {ws.title}")
            for row in data_rows:
                try:
                    if _row_empty(row):
                        continue
                    L = _norm(_cell(row, hdr_lc, "lang") or "pl")
                    pub = _norm(_cell(row, hdr_lc, "publish") or "true") in {"1","true","tak","yes","on","prawda"}
                    if not pub:
                        continue
                    raw_slug = _cell(row, hdr_lc, "slug")
                    key = _norm(_cell(row, hdr_lc, "slugkey") or "")
                    tpl = _cell(row, hdr_lc, "template") or "page.html"
                    parent = _cell(row, hdr_lc, "parentslug") or _cell(row, hdr_lc, "parent") or ""
                    order_v = _cell(row, hdr_lc, "order") or "999"

                    orig_L = L
                    slug_lang, rel = split_slug(raw_slug)
                    if slug_lang and slug_lang != orig_L:
                        report.append(
                            f"[warn] slug/lang mismatch: slug={raw_slug!r} col={orig_L} -> {slug_lang}"
                        )
                    L = slug_lang or orig_L
                    if slug_lang:
                        rel = rel
                    else:
                        rel = (raw_slug or "").strip("/")
                    canon_slug = f"/{L}/{rel}/"
                    chk_L, chk_rel = split_slug(canon_slug)
                    if chk_L != L or chk_rel != rel:
                        report.append(
                            f"[warn] bad slug {raw_slug!r} -> {canon_slug!r}"
                        )
                    if raw_slug and raw_slug != canon_slug:
                        report.append(
                            f"[warn] slug normalized: {raw_slug!r} -> {canon_slug!r}"
                        )
                    if not key:
                        key = (rel or "home") if rel else "home"
                    if rel == "home":
                        rel = ""
                    if not rel and key != "home":
                        rel = key

                    parent_key = (parent or "").strip()
                    if parent_key.startswith(f"/{L}/"):
                        parent_key = parent_key[len(f"/{L}/"):]
                    parent_key = parent_key.strip("/")

                    meta = {
                        "h1": _cell(row, hdr_lc, "h1"),
                        "title": _cell(row, hdr_lc, "title"),
                        "seo_title": _cell(row, hdr_lc, "seo_title"),
                        "meta_desc": _cell(row, hdr_lc, "meta_desc"),
                        "hero_alt": _cell(row, hdr_lc, "hero_alt"),
                        "hero_image": _cell(row, hdr_lc, "hero_image"),
                        "og_image": _cell(row, hdr_lc, "og_image"),
                        "canonical": _cell(row, hdr_lc, "canonical"),
                        "cta_label": _cell(row, hdr_lc, "cta_label"),
                        "cta_href": _cell(row, hdr_lc, "cta_href"),
                        "cta_phone": _cell(row, hdr_lc, "cta_phone"),
                        "whatsapp": _cell(row, hdr_lc, "whatsapp"),
                    }
                    meta_clean = {k: v for k, v in meta.items() if v}
                    for fld in ("h1", "title", "seo_title", "meta_desc"):
                        if not meta_clean.get(fld):
                            report.append(
                                f"[warn] missing {fld} for {L}/{key}"
                            )

                    pages_rows.append(
                        {
                            "lang": L,
                            "key": key,
                            "slug": rel,
                            "parent_key": parent_key,
                            "template": tpl,
                            "order": int(float(order_v or "999")),
                            "meta": meta_clean,
                        }
                    )
                    routes.setdefault(key, {})[L] = rel
                    pm = page_meta.setdefault(L, {}).setdefault(key, {})
                    pm.update(meta_clean)
                except IndexError:
                    continue

        if is_menu:
            report.append(f"[detect] menu-like: {ws.title}")
            for row in data_rows:
                try:
                    if _row_empty(row):
                        continue
                    L = _norm(_cell(row, hdr_lc, "lang") or "pl")
                    if not _norm(_cell(row, hdr_lc, "enabled") or "true") in {"1","true","tak","yes","on","prawda"}:
                        continue
                    label = _cell(row, hdr_lc, "label")
                    href = _cell(row, hdr_lc, "href")
                    if not label or not href:
                        continue
                    parent = _cell(row, hdr_lc, "parent") or ""
                    order_v = _cell(row, hdr_lc, "order") or "999"
                    col_v = _cell(row, hdr_lc, "col") or "1"
                    menu_rows.append(
                        {
                            "lang": L,
                            "label": label,
                            "href": href,
                            "parent": parent,
                            "order": int(float(order_v or "999")),
                            "col": int(float(col_v or "1")),
                            "enabled": True,
                        }
                    )
                except IndexError:
                    continue

        if is_meta:
            report.append(f"[detect] meta-like: {ws.title}")
            for row in data_rows:
                try:
                    if _row_empty(row):
                        continue
                    L = _norm(_cell(row, hdr_lc, "lang") or "pl")
                    key = _norm(_cell(row, hdr_lc, "key") or "")
                    if not key:
                        continue
                    pm = page_meta.setdefault(L, {}).setdefault(key, {})
                    for fld in ("title", "seo_title", "description", "og_image", "canonical"):
                        val = _cell(row, hdr_lc, fld)
                        if val:
                            pm[fld] = val
                    known = {"lang", "key", "title", "seo_title", "description", "og_image", "canonical"}
                    for h in hdr_lc:
                        if h in known:
                            continue
                        val = _cell(row, hdr_lc, h)
                        if val:
                            pm[h] = val
                except IndexError:
                    continue

        if is_blocks:
            report.append(f"[detect] blocks-like: {ws.title}")
            for row in data_rows:
                try:
                    if _row_empty(row):
                        continue
                    L = _norm(_cell(row, hdr_lc, "lang") or "pl")
                    path = _cell(row, hdr_lc, "path")
                    if not path:
                        key = _norm(_cell(row, hdr_lc, "key") or "")
                        section = _norm(_cell(row, hdr_lc, "section") or "")
                        if key and section:
                            path = f"pages/{key}/{section}"
                    if not path:
                        continue
                    b = blocks.setdefault(L, {}).setdefault(path.lstrip("/"), {})
                    html_val = _cell(row, hdr_lc, "html")
                    if html_val:
                        b["html"] = html_val
                    body_val = _cell(row, hdr_lc, "body")
                    if body_val and "html" not in b:
                        b["body"] = body_val
                    for fld in ("title", "cta_label", "cta_href"):
                        val = _cell(row, hdr_lc, fld)
                        if val:
                            b[fld] = val
                except IndexError:
                    continue

        if is_blog:
            report.append(f"[detect] blog-like: {ws.title}")
            for row in data_rows:
                if _row_empty(row):
                    continue
                rec: Dict[str, Any] = {}
                for idx, col_name in enumerate(hdr_lc):
                    v = row[idx] if idx < len(row) else ""
                    rec[col_name] = "" if v is None else str(v).strip()
                raw_slug = rec.get("slug") or ""
                L_blog = _norm(rec.get("lang") or "pl")
                slug_lang, rel = split_slug(raw_slug)
                if slug_lang and slug_lang != L_blog:
                    report.append(
                        f"[warn] blog slug/lang mismatch: slug={raw_slug!r} col={L_blog} -> {slug_lang}"
                    )
                canon_rel = rel if slug_lang else raw_slug.strip("/")
                canon_slug = f"/{L_blog}/{canon_rel}/"
                chk_L, chk_rel = split_slug(canon_slug)
                if chk_L != L_blog or chk_rel != canon_rel:
                    report.append(
                        f"[warn] bad blog slug {raw_slug!r} -> {canon_slug!r}"
                    )
                if raw_slug and raw_slug != canon_slug:
                    report.append(
                        f"[warn] blog slug normalized: {raw_slug!r} -> {canon_slug!r}"
                    )
                blog_rows.append(rec)

        if is_routes_sheet:
            report.append(f"[detect] routes-like: {ws.title}")
            langs = [h for h in hdr_lc if h in LOCALES]
            for row in data_rows:
                if _row_empty(row):
                    continue
                slug_key = _cell(row, hdr_lc, "slugkey")
                if not slug_key:
                    continue
                for L in langs:
                    idx = hdr_lc.index(L)
                    v = row[idx] if idx < len(row) else ""
                    if v:
                        routes.setdefault(slug_key, {})[L] = str(v).strip().lstrip("/").rstrip("/")

        if is_strings:
            report.append(f"[detect] strings-like: {ws.title}")
            langs = [h for h in hdr_lc if h in LOCALES]
            for row in data_rows:
                if _row_empty(row):
                    continue
                key = _cell(row, hdr_lc, "key")
                if not key:
                    continue
                rec = {"key": key}
                for L in langs:
                    idx = hdr_lc.index(L)
                    v = row[idx] if idx < len(row) else ""
                    if v:
                        rec[L] = str(v).strip()
                strings_rows.append(rec)

        if is_media:
            report.append(f"[detect] media-like: {ws.title}")
            for row in data_rows:
                if _row_empty(row):
                    continue
                rec = {}
                for idx, col_name in enumerate(hdr):
                    v = row[idx] if idx < len(row) else ""
                    rec[str(col_name).strip()] = "" if v is None else str(v).strip()
                media_rows.append(rec)

        if is_company:
            report.append(f"[detect] company-like: {ws.title}")
            for row in data_rows:
                if _row_empty(row):
                    continue
                rec = {}
                for idx, col_name in enumerate(hdr):
                    v = row[idx] if idx < len(row) else ""
                    rec[str(col_name).strip()] = "" if v is None else str(v).strip()
                company_rows.append(rec)

        if is_redirects:
            report.append(f"[detect] redirects-like: {ws.title}")
            for row in data_rows:
                if _row_empty(row):
                    continue
                src = _cell(row, hdr_lc, "from")
                dst = _cell(row, hdr_lc, "to")
                if src and dst:
                    redirect_rows.append({"from": src, "to": dst})

        if klass == "collection":
            it2 = ws.iter_rows(values_only=True)
            next(it2, None)
            for row in it2:
                if _row_empty(row):
                    continue
                vals = list(row or [])
                rec: Dict[str, str] = {}
                for idx, col_name in enumerate(hdr):
                    v = vals[idx] if idx < len(vals) else ""
                    rec[str(col_name).strip()] = "" if v is None else str(v).strip()

                lang = _norm(rec.get("lang") or "pl")
                collections.setdefault(ws.title, {}).setdefault(lang, []).append(rec)
                slug_lang, rel = split_slug(rec.get("slug") or "")
                if not slug_lang:
                    slug_lang = lang
                    rel = (rec.get("slug") or "").strip("/")
                key = _norm(rec.get("slugKey") or rec.get("slug") or "")
                if rel == "home":
                    rel = ""
                if not rel and key:
                    rel = key
                if key:
                    routes.setdefault(key, {})[slug_lang] = rel

    report.append(f"[rows] pages_rows={len(pages_rows)}, menu_rows={len(menu_rows)}")
    report.append(
        f"[result] meta_langs={len(page_meta)}, blocks_langs={len(blocks)}"
    )

    pages_by_key: Dict[str, Dict[str, Dict[str, Any]]] = {}
    for r in pages_rows:
        pages_by_key.setdefault(r["key"], {})[r["lang"]] = r

    for k, per_lang in routes.items():
        for L, rel in per_lang.items():
            if rel == "home":
                per_lang[L] = ""

    return {
        "menu_rows": menu_rows,
        "page_meta": page_meta,
        "blocks": blocks,
        "page_routes": routes,
        "routes": routes,
        "pages_rows": pages_rows,
        "pages_by_key": pages_by_key,
        "collections": collections,
        "blog_rows": blog_rows,
        "strings": strings_rows,
        "media": media_rows,
        "company": company_rows,
        "redirects": redirect_rows,
        "report": "\n".join(report),
    }

