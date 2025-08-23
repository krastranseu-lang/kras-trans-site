#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Kras-Trans • Static Builder (MAX+ • scalony)
Autor: Kras-Trans Project Builder

CELE:
- Połączenie pełnych funkcji starego buildera (P0–P20) z dodatkami „nowego patcha”.
- Zgodność z pages.yml v6, i18n (pl,en,de,fr,it,ru,ua), SEO, performance, a11y.

NAJWAŻNIEJSZE FUNKCJE:
- ENV override: SITE_URL, GA_ID, GSC_VERIFICATION, INDEXNOW_KEY, BING_SITE_AUTH_USER, NEWS_ENABLED
- CMS: wczytywanie z lokalnych plików data/cms, robust timeout/log
- Kopiowanie assets/ → dist/assets
- Render Jinja + autolinki + sanity DOM (a11y/perf)
- Head injections (jeśli brakuje w szablonie): GA (gtag), GSC meta, canonical, hreflang, OG/Twitter, JSON-LD
- Root "/" = redirect do /{defaultLang}/ + GSC meta + canonical
- City×Service generator + thin/noindex + OG-image (opcjonalnie)
- Sitemapy z alternates (xhtml:link), news-sitemap (okno godz.), robots.txt
- Redirect stubs, BingSiteAuth.xml, {INDEXNOW_KEY}.txt, /admin/indexing.html (Apps Script dispatch)
- On-site search index (per lang), RSS/Atom, link-checker, raporty

WYMAGANE PAKIETY:
  PyYAML, Jinja2, Markdown, requests, beautifulsoup4, lxml
OPCJONALNE:
  Pillow (OG-image), python-slugify (nieobowiązkowo)

UŻYCIE (CI):
  python -u tools/build.py
"""
import os, json, shutil
from pathlib import Path
from bs4 import BeautifulSoup
try:
    import cms_ingest  # nasz mały moduł do czytania XLSX (pkt 4 poniżej)
except Exception:
    cms_ingest = None
import re, io, csv, math, sys, time, glob, hashlib, unicodedata, pathlib
from datetime import datetime, timezone, timedelta
from typing import Dict, Any, List, Tuple, Iterable, Optional, Set

# --------------------------- ZALEŻNOŚCI ------------------------------------
try:
    import yaml
    from jinja2 import Environment, FileSystemLoader, select_autoescape, TemplateNotFound
    from markdown import markdown
    import requests
    import menu_builder  # tools/menu_builder.py
    try:
        from slugify import slugify as _slugify
    except Exception:
        _slugify = None
    try:
        from PIL import Image, ImageDraw, ImageFont
        PIL_OK = True
    except Exception:
        PIL_OK = False
except Exception as e:
    print("Brak wymaganych pakietów. Zainstaluj requirements.txt.", file=sys.stderr)
    raise

# --- SSR (HOME): składamy sekcje 1:1 z CMS na etapie builda ---------------
def _routes_map() -> Dict[str, Dict[str, str]]:
    """mapa slugKey -> { lang: '/{lang}/{slug}/' } na podstawie CMS.hreflang lub z Pages"""
    out: Dict[str, Dict[str, str]] = {}
    hre = CMS.get("hreflang") or {}
    for key, m in hre.items():
        out[key] = {}
        for L, href in (m or {}).items():
            # href to pełny URL — zamieniamy na ścieżkę
            try:
                path = "/" + href.split("://",1)[-1].split("/",1)[-1]  # po domenie
                out[key][L] = path if path.endswith("/") else (path + "/")
            except Exception:
                pass
    # fallback z pages, jeśli brak hreflang
    if not out:
        for p in CMS.get("pages", []):
            lang = (p.get("lang") or DEFAULT_LANG).lower()
            slugKey = p.get("slugKey") or (p.get("slug") or "")
            if not slugKey: slugKey = "home"
            out.setdefault(slugKey, {})
            out[slugKey][lang] = f"/{lang}/{(p.get('slug') or '') + '/' if p.get('slug') else ''}"
    return out

def _ssr_home(lang:str) -> Dict[str, Any]:
    """Zwraca gotowe sekcje HOME (hero/services/faq) do wstrzyknięcia w HTML."""
    L = (lang or DEFAULT_LANG).lower()
    pages = [dict(p) for p in CMS.get("pages", []) if (p.get("lang") or DEFAULT_LANG).lower()==L and p.get("publish", True)]
    if not pages:
        # fallback do defaultLang, jeśli dla danego języka jeszcze nie ma danych
        pages = [dict(p) for p in CMS.get("pages", []) if (p.get("lang") or DEFAULT_LANG).lower()==DEFAULT_LANG and p.get("publish", True)]
    strings = { (s.get("key") or s.get("Key") or "").strip(): s for s in CMS.get("strings", []) }
    STR = lambda key: (strings.get(key, {}).get(L) or strings.get(key, {}).get("pl") or "").strip()
    # HERO: rekord home
    home = next((p for p in pages if (p.get("slugKey") or p.get("slug") or "") in ("home","")), {})
    hero = {
        "title": home.get("h1") or home.get("title") or "",
        "lead":  home.get("lead") or "",
        "kpi":   [],  # opcjonalnie z Blocks/Strings
        "cta_primary":   {"label": home.get("cta_label") or STR("cta_quote_primary"),   "slugKey": "quote"},
        "cta_secondary": {"label": home.get("cta_secondary") or STR("cta_quote_secondary"), "slugKey": "contact"},
        "image": {"src": home.get("hero_image") or home.get("og_image") or "",
                  "srcset":"", "alt": home.get("hero_alt") or home.get("h1") or ""}
    }
    # SERVICES: wszystkie type=service
    svcs=[]
    for s in pages:
        if (s.get("type") or "").lower()!="service":
            continue
        svcs.append({
            "icon":"", "title": s.get("h1") or s.get("title") or "",
            "desc": s.get("lead") or "",
            "slugKey": s.get("slugKey") or (s.get("slug") or ""),
            "cta": {"label": STR("cta_quote_secondary") or ""}
        })
    svcs.sort(key=lambda x: (next((p.get("order",0) for p in pages if (p.get("slugKey")==x["slugKey"])), 0)))
    # FAQ: enabled + przypięte do home (page_slug='home' albo puste)
    faqs=[]
    for f in CMS.get("faq", []):
        fL = (f.get("lang") or L).lower()
        if fL!=L:
            continue
        enabled = str(f.get("enabled","true")).lower() not in ("false","0","no")
        if not enabled:
            continue
        pg = (f.get("page_slug") or f.get("slugKey") or "") or "home"
        if pg in ("", "home"):
            faqs.append({"q": f.get("q",""), "a": f.get("a","")})
    # Sekcyjne nagłówki z Strings (opcjonalnie)
    sect_titles = {
        "services": STR("services_h2"), "faq": STR("faq_h2"),
        "industries": STR("ind_h2"), "coverage": STR("cov_h2"),
        "process": STR("proc_h2"), "trust": STR("trust_h2"),
        "testimonials": STR("testi_h2"), "partners": STR("partners_h2"),
        "fleet": STR("fleet_h2"), "pricing": STR("pricing_h2"),
        "insights": STR("ins_h2")
    }
    sect_subs   = {
        "services": STR("services_sub"), "faq": STR("faq_sub"),
        "industries": STR("ind_sub"), "coverage": STR("cov_sub"),
        "process": STR("proc_sub"), "trust": STR("trust_sub"),
        "testimonials": STR("testi_sub"), "partners": STR("partners_sub"),
        "fleet": STR("fleet_sub"), "pricing": STR("pricing_sub"),
        "insights": STR("ins_sub")
    }
    return {
        "hero": hero,
        "services": svcs,
        "faq": faqs,
        "home": {"section_titles": sect_titles, "section_subtitles": sect_subs},
        "routes": _routes_map()
    }

# --------------------------- POMOCNICZE ------------------------------------
ROOT = Path(".")
DIST = Path("dist")
DIST.mkdir(parents=True, exist_ok=True)
DATA = Path("data")
OUT = DIST

UTC  = lambda dt=None: (dt or datetime.now(timezone.utc)).isoformat(timespec="seconds")

def read_yaml(path: "str|pathlib.Path") -> Dict[str, Any]:
    p = pathlib.Path(path)
    raw = p.read_text("utf-8")
    # Normalizacja, żeby YAML nie walił się na tabach/BOM/CRLF
    if raw.startswith("\ufeff"):  # BOM
        raw = raw.lstrip("\ufeff")
    raw = raw.replace("\r\n", "\n").replace("\r", "\n")
    raw = raw.replace("\t", "  ")  # TAB -> 2 spacje
    try:
        return yaml.safe_load(raw)
    except yaml.YAMLError as e:
        print("[YAML] Parse error:", e, file=sys.stderr)
        mark = getattr(e, "problem_mark", None)
        if mark:
            err_line = mark.line + 1
            start = max(1, err_line - 3)
            end = err_line + 3
            lines = raw.split("\n")
            for i in range(start, min(end, len(lines)) + 1):
                prefix = ">>" if i == err_line else "  "
                print(f"{prefix} {i:4d}: {lines[i-1]}", file=sys.stderr)
        raise

def read_text(p: pathlib.Path) -> str:
    return p.read_text("utf-8") if p.exists() else ""

def write_text(p: Path, s: str):
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(s, "utf-8")

write = write_text

def ensure_dir(p: pathlib.Path):
    p.mkdir(parents=True, exist_ok=True)

def norm_slug(s: str) -> str:
    s = unicodedata.normalize("NFD", s or "").encode("ascii","ignore").decode("ascii")
    s = re.sub(r"&","-and-", s.lower())
    s = re.sub(r"[^a-z0-9]+","-", s).strip("-")
    return re.sub(r"-{2,}","-", s)

def url_from(lang: str, slug: str) -> str:
    return f"/{lang}/{(slug + '/') if slug else ''}"

def canonical(site_url: str, lang: str, slug: str, canonical_path: Optional[str]=None) -> str:
    path = canonical_path or url_from(lang, slug)
    return site_url.rstrip("/") + path

def _canonical_url(base_url: str, current_path: str, override: str | None) -> str:
    """
    Zwraca pełny kanoniczny URL.
    - jeśli override to absolutny http(s) → zwraca override
    - jeśli override to ścieżka względna /foo/bar/ → base_url + override
    - jeśli override puste → base_url + current_path
    """
    base = (base_url or "").rstrip("/")
    if override:
        o = override.strip()
        if o.startswith("http://") or o.startswith("https://"):
            return o
        if not o.startswith("/"):
            o = "/" + o
        return base + o
    # fallback
    if not current_path.startswith("/"):
        current_path = "/" + current_path
    return base + current_path

def md_to_html(md: str) -> str:
    if not md: return ""
    return markdown(md, extensions=["extra","sane_lists","tables","toc"])

def soupify(html: str) -> BeautifulSoup:
    return BeautifulSoup(html or "", "lxml")

def text_of(html: str) -> str:
    return soupify(html).get_text(" ", strip=True)

def is_external(href: str, site_url: str) -> bool:
    return bool(re.match(r"^https?://", href)) and (not href.startswith(site_url))

def set_img_defaults(soup: BeautifulSoup):
    # lazy + hero LCP
    for img in soup.find_all("img"):
        if "loading" not in img.attrs:   img["loading"]="lazy"
        if "decoding" not in img.attrs:  img["decoding"]="async"
        if img.find_parent(id="hero") or "hero-media" in (img.get("class") or []):
            img["fetchpriority"]="high"; img["loading"]="eager"

def set_ext_link_attrs(soup: BeautifulSoup, site_url: str):
    for a in soup.find_all("a", href=True):
        href=a["href"]
        if href.startswith("mailto:") or href.startswith("tel:"): continue
        if is_external(href, site_url):
            rel=set(a.get("rel") or []); rel.update(["noopener","noreferrer"]); a["rel"]=list(rel)
            a["target"]="_blank"

def hash_stable(s: str) -> int:
    h=5381
    for ch in (s or ""): h=((h<<5)+h)+ord(ch)
    return abs(h)

def simhash(tokens: List[str], bits:int=64)->int:
    v=[0]*bits
    for t in tokens:
        h=int(hashlib.md5(t.encode("utf-8")).hexdigest(),16)
        for i in range(bits):
            v[i]+=1 if (h>>i)&1 else -1
    out=0
    for i in range(bits):
        if v[i]>0: out|=(1<<i)
    return out

def hamming(a:int,b:int)->int:
    return (a^b).bit_count()

def tfidf_keywords(text:str, top:int=8) -> List[str]:
    tokens=[w.lower() for w in re.findall(r"[a-zA-ZąćęłńóśźżĄĆĘŁŃÓŚŹŻ0-9]{3,}", text)]
    freq={}
    for t in tokens: freq[t]=freq.get(t,0)+1
    return [k for k,_ in sorted(freq.items(), key=lambda kv: kv[1], reverse=True)[:top]]

# --------------------------- KONFIG + ENV -----------------------------------
CFG = read_yaml("pages.yml")
C   = CFG.get("constants", {})

def _env(name: str, default: Any) -> Any:
    v = os.getenv(name)
    return default if v is None or str(v).strip()=="" else v

SITE_URL = (_env("SITE_URL", C.get("SITE_URL","")) or "").rstrip("/")
GA_ID    = _env("GA_ID", C.get("GA_ID",""))
GSC      = _env("GSC_VERIFICATION", C.get("GSC_VERIFICATION",""))
INDEXNOW_KEY = _env("INDEXNOW_KEY", C.get("INDEXNOW_KEY",""))
BING_USER    = _env("BING_SITE_AUTH_USER", C.get("BING_SITE_AUTH_USER",""))
NEWS_ENABLED = str(_env("NEWS_ENABLED", C.get("NEWS_ENABLED", False))).lower() in ("1","true","yes")

DEFAULT_LANG = CFG.get("site",{}).get("defaultLang","pl")
LOCALES      = list((CFG.get("site",{}).get("locales") or {}).keys()) or ["pl"]

# --------------------------- KOPIOWANIE ASSETS ------------------------------
ASSETS_DIR = pathlib.Path("assets")
if ASSETS_DIR.exists():
    shutil.copytree(ASSETS_DIR, OUT / "assets", dirs_exist_ok=True)

# ---------------------------- ŚRODOWISKO JINJA -----------------------------
TEMPLATES = Path(CFG["paths"]["src"]["templates"])
env = Environment(
    loader=FileSystemLoader(TEMPLATES),
    autoescape=select_autoescape(["html"])
)

def render_template(name: str, ctx: Dict[str, Any]) -> str:
    return env.get_template(name).render(**ctx)

# Globalne dane dostępne w szablonach
env.globals.update({
  "site": CFG.get("site", {}),
  "cms_endpoint": "",  # Apps Script wyłączony
  "ga_id": GA_ID,
  "gsc_verification": GSC,
  "assets": CFG.get("assets", {})
})

# Nawigacja + konfiguracja headera (partials/header.html)
env.globals.update({
    "nav": CFG.get("navigation", {}),
    "header_cfg": CFG.get("header", {})
})

# --------------------------- CMS: load (LOCAL) ------------------------------
def _cms_local_read() -> Dict[str, Any]:
    """
    Czyta CMS z data/cms/cms.json (preferowane), albo data/cms/cms.csv,
    albo data/cms/cms.xlsx. Zwraca {"ok": True, ...}.
    """
    base = pathlib.Path("data") / "cms"

    # JSON
    p_json = base / "cms.json"
    if p_json.exists():
        try:
            data = json.loads(p_json.read_text("utf-8"))
            if not data.get("ok"): data["ok"] = True
            print("[CMS] Lokalnie: data/cms/cms.json")
            return data
        except Exception as e:
            print(f"[CMS] Błąd JSON: {e}", file=sys.stderr)

    # CSV
    p_csv = base / "cms.csv"
    if p_csv.exists():
        try:
            rows = []
            raw = p_csv.read_text("utf-8")
            sep = "\t" if "\t" in raw else ("," if raw.count(",") > raw.count(";") else ";")
            for r in csv.DictReader(io.StringIO(raw), delimiter=sep):
                rows.append({(k or "").strip(): (v or "").strip() for k, v in r.items()})
            return {"ok": True, "rows": rows}
        except Exception as e:
            print(f"[CMS] Błąd CSV: {e}", file=sys.stderr)

    # XLSX
    env_path = os.environ.get("LOCAL_XLSX")
    p_xlsx = pathlib.Path(env_path) if env_path else (base / "cms.xlsx")
    if not p_xlsx.exists():
        p_xlsx = base / "cms.xlsx"
    if p_xlsx.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(p_xlsx, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            idx = {h: i for i, h in enumerate(headers)}
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = {h: (row[idx[h]] if h in idx else "") for h in headers}
                rows.append({(k or "").strip(): (str(v or "").strip()) for k, v in d.items()})
            print(f"[CMS] Lokalnie: {p_xlsx}")
            return {"ok": True, "rows": rows}
        except Exception as e:
            print(f"[CMS] Błąd XLSX: {e}", file=sys.stderr)

    # Minimalny fallback (strona główna)
    return {
        "ok": True,
        "pages": [{
            "lang": DEFAULT_LANG, "slugKey": "home", "slug": "",
            "type": "home",
            "h1": "Kras-Trans — transport i spedycja",
            "title": "Kras-Trans — transport i spedycja",
            "meta_desc": "Transport i spedycja w Polsce i UE.",
            "body_md": "## Start\n\nTreść domyślna (lokalny CMS nie wykrył pliku)."
        }],
        "hreflang": {}, "redirects": [], "blocks": [], "faq": []
    }

def load_cms() -> Dict[str, Any]:
    base = pathlib.Path("data") / "cms"
    if cms_ingest:
        try:
            data = cms_ingest.load_all(base)
            if data:
                print(data.get("report", ""))
                blocks_list = []
                for lang, m in (data.get("blocks") or {}).items():
                    for path, obj in m.items():
                        b = {"lang": lang, "path": path}
                        b.update(obj)
                        blocks_list.append(b)
                cms = {
                    "ok": True,
                    "pages": data.get("pages_rows", []),
                    "blocks": blocks_list,
                    "faq": data.get("faq_rows", []),
                    "strings": data.get("strings", []),
                    "props": data.get("props_rows", []),
                    "hreflang": data.get("page_routes", {}),
                    "menu_rows": data.get("menu_rows", []),
                    "page_meta": data.get("page_meta", {})
                }
                return cms
        except Exception as e:
            print(f"[CMS] cms_ingest error: {e}", file=sys.stderr)
    return _cms_local_read()

CMS = load_cms()

# ---------------------------- CSV: cities / keywords ------------------------
def read_csv(path:str, dialect="auto")->List[Dict[str,str]]:
    p=pathlib.Path(path)
    if not p.exists(): return []
    raw = p.read_text("utf-8")
    sep=";"
    if dialect=="auto":
        if "\t" in raw: sep="\t"
        elif raw.count(",")>raw.count(";"): sep=","
    rdr=csv.DictReader(io.StringIO(raw), delimiter=sep)
    rows=[]
    for row in rdr:
        rows.append({(k or "").strip(): (v or "").strip() for k,v in row.items()})
    return rows

def csv_map(rows:List[Dict[str,str]], mapping:Dict[str,List[str]])->List[Dict[str,str]]:
    out=[]
    alias_flat=set()
    for aliases in mapping.values(): alias_flat.update(aliases)
    for r in rows:
        o={}
        for canonical_name, aliases in mapping.items():
            v=""
            for a in aliases:
                if a in r and r[a]:
                    v=r[a]; break
            o[canonical_name]=v
        for k,v in r.items():
            if k not in alias_flat: o[k]=v
        out.append(o)
    return out

cities_src = CFG.get("sources",{}).get("cities_csv",{})
cities_rows = csv_map(
    read_csv(cities_src.get("path",""), cities_src.get("dialect","auto")),
    cities_src.get("map_columns",{
        "city":["city","miasto"],"voivodeship":["voivodeship","region"],"slug":["slug"],"lang":["lang"]
    })
) if cities_src else []

kw_src = CFG.get("sources",{}).get("keywords_csv",{})
kw_rows = csv_map(
    read_csv(kw_src.get("path",""), kw_src.get("dialect","auto")),
    kw_src.get("map_columns",{
        "lang":["lang"],"term":["term","keyword"],"type":["type"],"weight":["weight"],"anchor":["anchor"]
    })
) if kw_src else []

# ---------------------------- POMOC: wybór szablonu ------------------------
def choose_template(page:Dict[str,Any])->str:
    rules=CFG.get("template_rules",{})
    t=(page.get("type") or "").lower()
    if t and (t in (rules.get("by_type") or {})): return rules["by_type"][t]
    sk=page.get("slugKey") or ""
    if sk and (sk in (rules.get("by_slugKey") or {})): return rules["by_slugKey"][sk]
    sl=page.get("slug") or ""
    for r in (rules.get("by_slug_pattern") or []):
        if re.search(r["match"], sl): return r["template"]
    return CFG["templates"]["page"]

# ---------------------------- PAGES: bazowe --------------------------------
def base_pages() -> List[Dict[str, Any]]:
    pages = []
    for p in CMS.get("pages", []):
        lang = p.get("lang") or DEFAULT_LANG
        slug = p.get("slug", "")
        ctx = dict(p)
        ctx["canonical"] = canonical(SITE_URL, lang, slug, p.get("canonical_path"))
        ctx["og_image"] = p.get("og_image") or CFG.get("seo", {}).get("open_graph", {}).get("default_image")
        ctx["body_html"] = p.get("body_html") or md_to_html(p.get("body_md", ""))

        if not ctx.get("seo_title"):
            ctx["seo_title"] = (p.get("seo_title") or p.get("title") or p.get("h1") or "")

        if not ctx.get("title"):
            ctx["title"] = (p.get("h1") or ctx["seo_title"])

        ctx["template"] = choose_template(ctx)
        ctx["__from"] = p.get("__from", "pages")   # ← TEN wiersz zostaje
        pages.append(ctx)

    return pages   # ← poza pętlą!

# ------------------------ PAGES: city × service ----------------------------
def merge_places()->List[Dict[str,str]]:
    out=[]; seen=set()
    for row in CMS.get("places", []):
        lang=(row.get("lang") or "pl").lower()
        city=row.get("city") or ""
        slug=row.get("slug") or norm_slug(city)
        key=f"{lang}::{slug}"
        if key in seen: continue
        seen.add(key)
        out.append({"lang":lang,"city":city,"voivodeship":row.get("voivodeship") or row.get("region") or "","slug":slug})
    for row in cities_rows:
        lang=(row.get("lang") or "pl").lower()
        city=row.get("city") or ""
        slug=row.get("slug") or norm_slug(city)
        key=f"{lang}::{slug}"
        if key in seen: continue
        seen.add(key)
        out.append({"lang":lang,"city":city,"voivodeship":row.get("voivodeship") or "","slug":slug})
    return out

def generate_city_service()->List[Dict[str,Any]]:
    cfg=CFG.get("collections",{}).get("city_service",{})
    if not cfg: return []
    per_lang = cfg.get("limits",{}).get("perLang", 0)
    max_total= cfg.get("limits",{}).get("maxPages", 0)
    langs    = cfg.get("langs", LOCALES)
    services=[p for p in CMS.get("pages",[]) if (p.get("type")=="service" and p.get("publish",True))]
    places=merge_places()
    out=[]; total=0
    for L in langs:
        used=0
        svcL=[s for s in services if (s.get("lang") or L)==L]
        for svc in svcL:
            svc_slug=svc.get("slug") or norm_slug(svc.get("slugKey") or svc.get("h1") or "service")
            svc_h1=svc.get("h1") or svc.get("title") or svc_slug
            for city in places:
                if (city.get("lang") or L).lower()!=L: continue
                city_slug=city.get("slug") or norm_slug(city.get("city") or "")
                if not city_slug: continue
                slug=norm_slug(f"{svc_slug}-{city_slug}")
                h1=f"{svc_h1} — {city.get('city')}"
                meta=f"Transport i spedycja — {svc_h1} w {city.get('city')}. Wycena w 15 min, kontakt 24/7."
                row={
                    "lang":L,"type":"city_service","slugKey":f"{svc.get('slugKey','service')}__{city_slug}",
                    "slug":slug,"template":choose_template({"type":"city_service","slug":slug}),
                    "publish":True,"h1":h1,"title":h1,"seo_title":f"{h1} | Kras-Trans","meta_desc":meta,
                    "hero_alt":h1,"lead":svc.get("lead") or svc.get("title") or "",
                    "og_image":svc.get("og_image") or CFG.get("seo",{}).get("open_graph",{}).get("default_image"),
                    "canonical_path":f"/{L}/{slug}/",
                    "city":city.get("city"),"voivodeship":city.get("voivodeship") or city.get("region") or "",
                    "service_slug":svc_slug,"service_h1":svc_h1,"__from":"city_service",
                    "body_html":""
                }
                out.append(row); used+=1; total+=1
                if used>=per_lang or total>=max_total: break
            if used>=per_lang or total>=max_total: break
        if total>=max_total: break
    return out

# ------------------------------ SEO / GATES --------------------------------
TITLE_MIN = CFG.get("seo",{}).get("titles",{}).get("min", 30)
TITLE_MAX = CFG.get("seo",{}).get("titles",{}).get("max", 65)
DESC_MIN  = CFG.get("seo",{}).get("descriptions",{}).get("min", 80)
DESC_MAX  = CFG.get("seo",{}).get("descriptions",{}).get("max", 165)
THIN_MIN_CHARS = CFG.get("collections",{}).get("city_service",{}).get("quality",{}).get("thin_min_chars", 400)

def clamp_len(s:str, minL:int, maxL:int)->Tuple[str,List[str]]:
    s=s or ""; warns=[]
    if len(s)<minL: warns.append(f"<{minL}")
    if len(s)>maxL:
        s=s[:maxL].rstrip()
        if not s.endswith("…"): s+="…"
        warns.append(f">{maxL}")
    return s, warns

def apply_quality(page:Dict[str,Any])->Tuple[Dict[str,Any],List[str]]:
    warns=[]
    page["seo_title"], w = clamp_len(page.get("seo_title") or page.get("title") or "", TITLE_MIN, TITLE_MAX); warns+=["title"+x for x in w]
    fallback_desc = text_of(page.get("body_html",""))[:180]
    page["meta_desc"],  w = clamp_len(page.get("meta_desc") or fallback_desc, DESC_MIN, DESC_MAX); warns+=["desc"+x for x in w]
    if not (page.get("h1") or "").strip(): warns.append("missing_h1")
    # thin content – tylko generator
    text_len=len(text_of(page.get("body_html","")))
    if page.get("__from")=="city_service" and text_len<THIN_MIN_CHARS:
        page["noindex"]=True; warns.append(f"thin({text_len})")
    return page, warns

# ------------------------------ JSON-LD ------------------------------------
def jsonld_blocks(page:Dict[str,Any])->List[Dict[str,Any]]:
    ld=[]
    ld.append({"@context":"https://schema.org","@type":"Organization","name":CFG.get("site",{}).get("brand","Kras-Trans"),"url":SITE_URL})
    if (page.get("slugKey")=="home" or (page.get("slug") or "")==""):
        ld.append({"@context":"https://schema.org","@type":"WebSite","name":CFG.get("site",{}).get("brand","Kras-Trans"),
                   "url": SITE_URL + f"/{page.get('lang', DEFAULT_LANG)}/",
                   "potentialAction":{"@type":"SearchAction","target": SITE_URL + "/search?q={query}","query-input":"required name=query"}})
    ld.append({"@context":"https://schema.org","@type":"BreadcrumbList","itemListElement":[
        {"@type":"ListItem","position":1,"name":CFG.get("site",{}).get("brand","Kras-Trans"),"item":SITE_URL},
        {"@type":"ListItem","position":2,"name":page.get("h1") or page.get("title") or page.get("slug") or "Page","item":page.get("canonical")}
    ]})
    # FAQ (po slugKey/slug)
    fq=[]
    lang=page.get("lang")
    need_keys = ((page.get("slugKey") or "").lower(), (page.get("slug") or "").lower())
    for f in CMS.get("faq", []):
        lang_f=(f.get("lang") or lang).lower()
        key=(f.get("page_slug") or f.get("slugKey") or "").lower()
        if lang_f==lang and key in need_keys:
            fq.append({"@type":"Question","name":f.get("q",""),"acceptedAnswer":{"@type":"Answer","text":f.get("a","")}})
    if fq: ld.append({"@context":"https://schema.org","@type":"FAQPage","mainEntity":fq[:30]})
    # typowe typy
    t=(page.get("type") or "page").lower()
    if t=="service":
        ld.append({"@context":"https://schema.org","@type":"Service","name":page.get("h1") or page.get("title")})
    if t=="blog_post":
        ld.append({"@context":"https://schema.org","@type":"Article","headline":page.get("h1") or page.get("title"),"mainEntityOfPage":page.get("canonical")})
    if t in ("job","job_post","jobposting"):
        ld.append({"@context":"https://schema.org","@type":"JobPosting","title":page.get("h1") or page.get("title"),
                   "hiringOrganization":{"@type":"Organization","name":CFG.get("site",{}).get("brand","Kras-Trans")}})
    if t=="city_service":
        company = (CMS.get("company") or [{}])[0]
        ld.append({"@context":"https://schema.org","@type":"LocalBusiness",
                   "name": CFG.get("site",{}).get("brand","Kras-Trans")+" - "+(page.get("h1") or ""),
                   "address":{"@type":"PostalAddress","addressLocality":page.get("city",""),"addressCountry":"PL"},
                   "telephone": company.get("telephone") or CFG.get("site",{}).get("contacts",{}).get("phone",""),
                   "url": page.get("canonical")})
    return ld

# ------------------------------ AUTOLINKI ----------------------------------
def fetch_explainer(lang:str)->str:
    blocks = [b for b in CMS.get("blocks", []) if str(b.get("enabled","true")).lower() not in ("false","0","no") and (b.get("type") or b.get("block"))=="explainer"]
    cand=[b for b in blocks if (b.get("lang") or lang).lower()==lang]
    if not cand: cand=blocks
    if not cand: return ""
    pick=cand[hash_stable(lang) % len(cand)]
    body = pick.get("body_md") or pick.get("desc") or pick.get("title") or ""
    return text_of(md_to_html(body))[:600]

def inject_autolinks(html:str, lang:str)->Tuple[str,int,int]:
    rules=[]
    for r in CMS.get("autolinks", []):
        if str(r.get("enabled","true")).lower() in ("false","0","no"): continue
        rules.append({"anchor":(r.get("anchor") or r.get("a") or "").strip(),
                      "href":(r.get("href") or "").strip(),
                      "lang":(r.get("lang") or lang).lower()})
    rules=[r for r in rules if r["anchor"] and r["href"] and r["lang"]==lang]
    if not rules: return html,0,0

    cfg=CFG.get("autolinks",{
        "inline":{"ignoreSelectors":["h1","h2","h3",".hero",".cta",".btn","nav","header","footer",".no-autolink"],"maxPerPage":6,"minDistanceChars":40},
        "fallback":{"enabled":True,"limit":3}
    })
    soup=soupify(html)
    ignore_nodes=set()
    for sel in cfg["inline"]["ignoreSelectors"]:
        for el in soup.select(sel): ignore_nodes.add(el)

    made=0; missed=[]
    pattern_cache={}
    maxN=cfg["inline"]["maxPerPage"]; minDist=cfg["inline"]["minDistanceChars"]

    for r in rules:
        if made>=maxN: break
        anchor=r["anchor"]; href=r["href"]
        found=False
        for p in soup.find_all(["p","li"]):
            if any(p is n or n in p.parents for n in ignore_nodes): continue
            if p.find("a"): continue
            txt=p.get_text()
            if len(txt)>(minDist*2):
                m=re.search(re.escape(anchor), txt, flags=re.IGNORECASE)
                if m and m.start()<minDist: continue
            if anchor not in pattern_cache:
                pattern_cache[anchor]=re.compile(rf"(?<!\w){re.escape(anchor)}(?!\w)", flags=re.IGNORECASE)
            html_p=str(p)
            new=pattern_cache[anchor].sub(f'<a href="{href}" title="{anchor}">{anchor}</a>', html_p, count=1)
            if new!=html_p:
                p.replace_with(soupify(new)); made+=1; found=True; break
        if not found: missed.append(r)

    fb=0
    if cfg["fallback"]["enabled"] and missed:
        cont = soup.find(id="content") or soup.find("article") or soup.body
        if cont:
            title_map={"pl":"Zobacz też","en":"See also","de":"Siehe auch","fr":"Voir aussi","it":"Vedi anche","ru":"См. также","ua":"Див. також"}
            h=soup.new_tag("h3"); h.string = title_map.get(lang,"See also"); cont.append(h)
            wrap=soup.new_tag("div", **{"class":"see-also cards"})
            expl = fetch_explainer(lang)
            for r in missed[:max(0, cfg["fallback"]["limit"]-(made))]:
                card=soup.new_tag("article", **{"class":"tile3d"})
                p=soup.new_tag("p"); p.string = expl or r["anchor"]
                a=soup.new_tag("a", href=r["href"], **{"class":"tile-link"}); a.string=r["anchor"]
                card.append(p); card.append(a); wrap.append(card); fb+=1
            cont.append(wrap)

    set_ext_link_attrs(soup, SITE_URL); set_img_defaults(soup)
    return str(soup), made, fb

# ------------------------------ OG IMAGE (opcja) ---------------------------
def og_image_for(page:Dict[str,Any])->Optional[str]:
    if not PIL_OK: return None
    t=(page.get("type") or "page").lower()
    if t not in ("service","city_service","page"): return None
    if page.get("og_image"): return page["og_image"]
    # generator PNG 1200x630
    w,h=1200,630
    img=Image.new("RGB",(w,h),(11,18,32))
    draw=ImageDraw.Draw(img)
    title=(page.get("h1") or page.get("title") or CFG.get("site",{}).get("brand","Kras-Trans"))[:90]
    sub=(page.get("meta_desc") or "")[:120]
    try:
        font1=ImageFont.truetype("DejaVuSans-Bold.ttf", 60)
        font2=ImageFont.truetype("DejaVuSans.ttf", 32)
    except Exception:
        font1=ImageFont.load_default(); font2=ImageFont.load_default()
    draw.text((60,220), title, fill=(233,237,246), font=font1)
    draw.text((60,320), sub,   fill=(168,179,199), font=font2)
    draw.rectangle([(0,h-10),(w,h)], fill=(34,195,166))
    out=OUT/"og"; out.mkdir(parents=True, exist_ok=True)
    name=f"{norm_slug(page.get('slugKey') or page.get('slug') or 'page')}.png"
    path=out/name; img.save(path, "PNG")
    return f"/og/{name}"

# ------------------------------ HEAD INJECTIONS -----------------------------
def ensure_head_injections(soup:BeautifulSoup, page:Dict[str,Any], hreflang_map:Dict[str,Dict[str,str]]):
    """
    Wstrzykuje: canonical, hreflang, OG/Twitter, GSC, GA, JSON-LD (jeśli brak).
    Bez duplikacji. Szanuje istniejące tagi z szablonów.
    """
    head = soup.find("head")
    if not head:
        head = soup.new_tag("head")
        if soup.html:
            soup.html.insert(0, head)
        else:
            soup.insert(0, head)

    def has_selector(sel:str)->bool:
        return bool(head.select_one(sel))

    def add_meta(**attrs):
        if not head.find("meta", attrs=attrs):
            head.append(soup.new_tag("meta", **attrs))

    def add_link(**attrs):
        # unikaj duplikatów
        for link in head.find_all("link"):
            if all(link.get(k) == v for k, v in attrs.items()):
                return
        head.append(soup.new_tag("link", **attrs))

    # canonical
    if page.get("canonical") and not has_selector('link[rel="canonical"]'):
        add_link(rel="canonical", href=page["canonical"])

    # hreflang alternates
    alts = (hreflang_map.get(page.get("slugKey","home"), {}) or {})
    for L, href in alts.items():
        add_link(rel="alternate", hreflang=L, href=href)

    # description
    if (page.get("meta_desc") or "") and not head.find("meta", attrs={"name":"description"}):
        add_meta(name="description", content=page["meta_desc"])
        # menu-bundle-version (SWR)
        if page.get("menu_version") and not head.find("meta", attrs={"name":"menu-bundle-version"}):
            add_meta(name="menu-bundle-version", content=str(page["menu_version"]))

    # OG/Twitter
    og_map = {
        "og:title": page.get("seo_title") or page.get("title") or page.get("h1") or "",
        "og:description": page.get("meta_desc") or "",
        "og:url": page.get("canonical") or "",
        "og:image": page.get("og_image") or CFG.get("seo",{}).get("open_graph",{}).get("default_image",""),
        "og:type": "website" if (page.get("slug","")=="" or page.get("type") in (None,"home","page")) else "article"
    }
    for prop,val in og_map.items():
        if val and not head.find("meta", attrs={"property":prop}):
            head.append(soup.new_tag("meta", **{"property":prop, "content":val}))
    if og_map["og:image"] and not head.find("meta", attrs={"name":"twitter:card"}):
        add_meta(name="twitter:card", content="summary_large_image")
    if og_map["og:title"] and not head.find("meta", attrs={"name":"twitter:title"}):
        add_meta(name="twitter:title", content=og_map["og:title"])
    if og_map["og:description"] and not head.find("meta", attrs={"name":"twitter:description"}):
        add_meta(name="twitter:description", content=og_map["og:description"])
    if og_map["og:image"] and not head.find("meta", attrs={"name":"twitter:image"}):
        add_meta(name="twitter:image", content=og_map["og:image"])

    # GSC verification
    if (GSC or "").strip() and not head.find("meta", attrs={"name":"google-site-verification"}):
        add_meta(name="google-site-verification", content=GSC)

    # GA (gtag) – wstrzykuj tylko, jeśli w całym dokumencie nie ma już configu
    if (GA_ID or "").strip():
        has_gtm_any = bool(soup.find("script", src=re.compile(r"googletagmanager\.com/gtag/js")))
        has_conf_any = bool(soup.find("script", string=re.compile(r"gtag\('config',\s*['\"]"+re.escape(GA_ID))))
        if not has_gtm_any:
            s = soup.new_tag("script", src=f"https://www.googletagmanager.com/gtag/js?id={GA_ID}")
            s["async"] = "async"
            head.append(s)
        if not has_conf_any:
            conf = soup.new_tag("script")
            conf.string = (
                "window.dataLayer=window.dataLayer||[];"
                "function gtag(){dataLayer.push(arguments);}gtag('js',new Date());"
                f"gtag('config','{GA_ID}',{{anonymize_ip:true}});"
            )
            head.append(conf)

    # JSON-LD – dołóż tylko jeśli w całym dokumencie nie ma ld+json
    if not soup.find("script", attrs={"type":"application/ld+json"}):
        try:
            ld = json.dumps(jsonld_blocks(page), ensure_ascii=False)
            ld_tag = soup.new_tag("script", type="application/ld+json")
            ld_tag.string = ld
            head.append(ld_tag)
        except Exception:
            pass

# --------- LINK GRAPH (pozostawione jak w starym; może być użyte w szabl.) --
def neighbors_for(
    city_pages: List[Dict[str, Any]],
    page: Dict[str, Any],
    k_region: int,
    k_alt: int
) -> List[Dict[str, Any]]:
    lang   = page.get("lang")
    region = (page.get("voivodeship") or "").strip().lower()
    city   = (page.get("city") or "").strip().lower()
    svc    = page.get("service_h1")

    # Strony z tym samym regionem, innym miastem, tą samą usługą
    same_region = [
        p for p in city_pages
        if (p.get("lang") == lang)
        and ((p.get("voivodeship") or "").strip().lower() == region)
        and ((p.get("city") or "").strip().lower() != city)
        and (p.get("service_h1") == svc)
    ]

    # Strony z tym samym miastem, inną usługą
    alt_service = [
        p for p in city_pages
        if (p.get("lang") == lang)
        and ((p.get("city") or "").strip().lower() == city)
        and (p.get("service_h1") != svc)
    ]

    out = same_region[:k_region] + alt_service[:k_alt]
    return out

# ------------------------------ RENDER / BUILD ------------------------------
def build_all():
    site = {
        "default_lang": CFG.get("default_lang") or CFG.get("site", {}).get("defaultLang", "pl"),
        "languages": CFG.get("languages") or LOCALES,
        "autogenerate_from_cms": CFG.get("autogenerate_from_cms", False),
    }
    languages = site.get("languages", ["pl"])
    dlang = site.get("default_lang", "pl")
    page_defs = CFG.get("pages", [])
    # === CMS: wczytaj XLSX z runnera (LOCAL_XLSX/CMS_SOURCE albo stała ścieżka) ===
    src_path = os.getenv("LOCAL_XLSX") or os.getenv("CMS_SOURCE") or "/Users/illia/Desktop/Kras_transStrona/CMS.xlsx"
    cms = {"menu_rows": [], "page_meta": {}, "blocks": {}, "report": "[cms] no module"}
    if cms_ingest:
        cms = cms_ingest.load_all((DATA / "cms"), explicit_src=Path(src_path))
        print(cms.get("report", "[cms] no report"))
    else:
        print("[cms] cms_ingest not available")
    global CMS
    CMS = cms
    # === Languages union from config + CMS ===
    langs_from_cms = sorted({r.get("lang", "pl") for r in (cms.get("pages_rows") or [])})
    languages = sorted(set(languages) | set(langs_from_cms))
    site["languages"] = languages
    slugs = cms.get("page_routes") or {}
    rows   = cms.get("pages_rows")  or []
    by_key_lang = {}
    for r in rows:
        by_key_lang[(r["key"], r["lang"])] = r

    def meta_get(L, K):
        return (cms.get("page_meta", {}).get(L, {}).get(K, {})) or {}

    # === Auto-pages z CMS: dopisz brakujące strony ===
    page_list = locals().get("page_defs") or locals().get("pages") or []
    existing  = {p.get("key") for p in page_list if isinstance(p, dict)}
    dlang     = site.get("default_lang", "pl")
    for key, per_lang in slugs.items():
        if key in existing:
            continue
        slug_map = {L: per_lang.get(L, "") for L in languages}
        row_dl = by_key_lang.get((key, dlang), {}) or {}
        tpl    = row_dl.get("template") or "page.html"
        parent = row_dl.get("parent_key") or "home"

        title_map = {}
        desc_map  = {}
        for L in languages:
            m = meta_get(L, key)
            title_map[L] = m.get("seo_title") or m.get("title") or key
            desc_map[L]  = m.get("description") or m.get("meta_desc") or ""

        page_list.append({
          "key": key,
          "template": tpl,
          "parent": parent,
          "slugs": slug_map,                      # <— kluczowe: routing z arkusza
          "title": title_map,
          "description": desc_map,
          "og_image": meta_get(dlang, key).get("og_image") or "/assets/img/placeholder.svg"
        })

    if "page_defs" in locals(): page_defs = page_list
    if "pages"     in locals(): pages     = page_list
    # --- REBUILD ROUTING PO AUTOGENIE ---
    _build_pages = page_list  # to jest nasza aktualna lista stron
    slugs = { p["key"]: p["slugs"] for p in _build_pages }   # <-- odświeżone slugs
    print(f"[cms] pages autogen: total_keys={len(slugs)}; after_merge={len(_build_pages)}")

    # helper path_for korzystający z aktualnych slugs
    def path_for(key: str, lang: str) -> str:
        try:
            s = slugs[key][lang]
        except Exception:
            s = ""
        return f"/{lang}/" if not s else f"/{lang}/{s}/"

    # jeżeli wcześniej wyliczałeś nav_urls przed autogenem – wylicz JESZCZE RAZ teraz:
    nav_keys = ["home","services","fleet","about","contact"] if "nav_keys" not in locals() else nav_keys
    nav_urls = { L: {k: path_for(k, L) for k in nav_keys if k in slugs} for L in languages }

    # === MENU: jeśli są wiersze z arkusza → buduj bundlery + HTML do SSR ===
    rows = cms.get("menu_rows") or []
    if not rows:
        def _menu_from_pages(pages_rows):
            out=[]
            titles={ (r.get("lang"), r.get("key")): (r.get("meta",{}).get("title") or r.get("key")) for r in pages_rows }
            for r in pages_rows:
                lang=r.get("lang")
                key=r.get("key")
                slug=r.get("slug") or ""
                parent=r.get("parent_key") or ""
                label=titles.get((lang,key), key)
                href=f"/{lang}/{slug + '/' if slug else ''}"
                item={"lang":lang,"label":label,"href":href,"parent":"","order":r.get("order",999),"col":1,"enabled":True}
                if parent:
                    parent_label=titles.get((lang,parent),"")
                    item["parent"]=parent_label
                out.append(item)
            return out
        rows = _menu_from_pages(cms.get("pages_rows", []))
        cms["menu_rows"] = rows
        print("[cms] menu_rows empty → built from pages_rows")
    if rows:
        bundles, html_by_lang = {}, {}
        for L in languages:
            b = menu_builder.build_bundle_for_lang(rows, L)
            bundles[L] = b
            html_by_lang[L] = menu_builder.render_nav_html(b)
        # zapisz bundlery do nowej i legacy ścieżki (żeby nie było 404)
        out_new = DIST / "assets" / "data" / "menu"
        out_old = DIST / "assets" / "nav"
        out_new.mkdir(parents=True, exist_ok=True)
        out_old.mkdir(parents=True, exist_ok=True)
        for L, b in bundles.items():
            p = out_new / f"bundle_{L}.json"
            p.write_text(json.dumps(b, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
            shutil.copy2(p, out_old / p.name)
    else:
        bundles, html_by_lang = {}, {}
        print("[cms] menu_rows empty → pozostaje dotychczasowe menu (jeśli jest)")
    CMS["pages"] = page_defs + CMS.get("pages", [])
    # === Wstrzyknięcie bloków (SSR) do elementów z data-api ===
    def _inject_blocks(html, lang):
        bl = cms.get("blocks", {}).get(lang, {})
        if not bl:
            return html
        soup = BeautifulSoup(html, "html.parser")
        for el in soup.select("[data-api]"):
            path = el.get("data-api","").split("?")[0].lstrip("/").rstrip("/")
            blk = bl.get(path)
            if not blk:
                continue
            if blk.get("html"):
                el.clear()
                el.append(BeautifulSoup(blk["html"], "html.parser"))
            else:
                if blk.get("title"):
                    h = el.find(["h1","h2","h3"])
                    if h: h.string = blk["title"]
                if blk.get("body"):
                    tgt = el.find(class_="lead") or el.find("p")
                    if tgt:
                        tgt.clear()
                        tgt.append(BeautifulSoup(blk["body"], "html.parser"))
                if blk.get("cta_label"):
                    btn = el.find("a", class_="btn-cta") or el.find("button", class_="btn-cta")
                    if btn:
                        btn.string = blk["cta_label"]
                        if btn.name == "a" and blk.get("cta_href"):
                            btn["href"] = blk["cta_href"]
        return str(soup)
    # === sanity: musi istnieć bundle dla domyślnego języka
    dlang_check = site.get("default_lang", "pl")
    assert (DIST/"assets"/"data"/"menu"/f"bundle_{dlang_check}.json").exists() or \
           (DIST/"assets"/"nav"/f"bundle_{dlang_check}.json").exists(), "❌ Brak bundla menu (404)"
    pages = base_pages()
    city  = generate_city_service()
    _build_pages = pages + city

    hreflang_map = CMS.get("hreflang", {})
    indexables: List[Tuple[str,str,str]] = []
    today=UTC()
    logs=[]
    autolink_inline=0; autolink_fb=0

    # simhash (P4) – wykrywanie duplikatów
    sim_index={}
    dup_warns=0

    # TF-IDF (P3) – z tekstu strony
    tfidf_map={}

    for p in _build_pages:
        lang = p.get("lang") or DEFAULT_LANG
        slug = p.get("slug", "")
        key = p.get("slugKey") or (p.get("slug") or "")
        if "key" not in p:
            p["key"] = key
        if "tsiny" in (p.get("key") or ""):
            print("[cms] route check:", {L: path_for(p["key"], L) for L in languages})

        # OG image generator (opcjonalnie; tylko gdy brak)
        gen_og = og_image_for(p)
        if gen_og:
            p["og_image"] = gen_og

        meta_title = p.get("seo_title") or ""
        meta_desc = p.get("meta_desc") or ""
        og_image = p.get("og_image") or ""

        # current path z routera
        try:
            current_path = path_for(key, lang)
        except Exception:
            try:
                current_path = "/" + lang + "/" + (slugs[key][lang] or "")
                if not current_path.endswith("/"):
                    current_path += "/"
            except Exception:
                current_path = "/" + lang + "/"

        # default canonical na bazie current_path
        canonical_url = _canonical_url(site.get("base_url", ""), current_path, None)

        # META override z CMS
        over = cms.get("page_meta", {}).get(lang, {}).get(key, {}) or {}
        if over.get("seo_title"):
            meta_title = over["seo_title"]
        if over.get("description"):
            meta_desc = over["description"]
        if over.get("og_image"):
            og_image = over["og_image"]
        # >>> USTAW canonical BEZPIECZNIE <<<
        canonical_url = _canonical_url(site.get("base_url", ""), current_path,
                                       over.get("canonical") or over.get("canonical_path"))

        cta_label = over.get("cta_label")
        cta_href = over.get("cta_href")

        p["canonical"] = canonical_url
        p["seo_title"] = meta_title
        p["meta_desc"] = meta_desc
        p["og_image"] = og_image

        # SEO gates
        p, warns = apply_quality(p)

        # JSON-LD (przekazany do Jinja, a dodatkowo do-inject po renderze jeśli brak)
        ld_html = '<script type="application/ld+json">' + json.dumps(jsonld_blocks(p), ensure_ascii=False) + '</script>'

        # SSR HOME: jeśli to strona "home", przekaż gotowe sekcje do szablonu
        ssr_ctx = _ssr_home(lang) if (p.get("type") or "").lower() == "home" else None

        # Render Jinja
        ctx = {
          "page": {**p, "menu_version": (bundles.get(lang, {}) or {}).get("version","")},
          "canonical": canonical_url,
          "hreflang": hreflang_map.get(p.get("slugKey","home"), {}),
          "ctas": {
            "primary": p.get("cta_label") or "Wycena transportu",
            "secondary": p.get("cta_secondary","")
          },
          "jsonld": ld_html,
          "ga_id": GA_ID,
          "gsc_verification": GSC,
          "ssr": ssr_ctx,
          # MENU (SSR + SWR)
          "menu_html": html_by_lang.get(lang, ""),
          "menu_bundle_inline": json.dumps(bundles.get(lang, {}), ensure_ascii=False),
          "menu_version": (bundles.get(lang, {}) or {}).get("version","")
        }
        tpl = (p.get("template") or "page.html").strip()
        candidate1 = f"pages/{tpl}"
        candidate2 = f"{tpl}"
        template_rel = candidate1 if (TEMPLATES / candidate1).exists() else candidate2
        try:
            page_html = render_template(template_rel, ctx)
        except TemplateNotFound:
            page_html = (
                f"<!doctype html><html lang=\"{lang}\"><head>"
                f"<meta charset='utf-8'><title>{p.get('seo_title','')}</title>"
                f"</head><body><h1>{p.get('h1','')}</h1></body></html>"
            )
        page_html = _inject_blocks(page_html, lang)
        if cta_label or cta_href:
            soup_cta = BeautifulSoup(page_html, "html.parser")
            btn = soup_cta.select_one(".btn-cta")
            if btn:
                if cta_label:
                    btn.string = cta_label
                if cta_href and btn.name == "a":
                    btn["href"] = cta_href
            page_html = str(soup_cta)
        # debug
        write(DIST/"_debug_cms.json", json.dumps({
          "menu_rows": len(cms.get("menu_rows",[])),
          "meta_langs": list(cms.get("page_meta",{}).keys()),
          "blocks_langs": list(cms.get("blocks",{}).keys())
        }, ensure_ascii=False, indent=2))

        # AUTOLINKI + fallback
        page_html, ai, fb = inject_autolinks(page_html, lang)
        autolink_inline += ai; autolink_fb += fb

        # OSTATECZNY DOM
        soup = soupify(page_html)
        set_ext_link_attrs(soup, SITE_URL); set_img_defaults(soup)
        ensure_head_injections(soup, ctx["page"], hreflang_map)

        # TF-IDF prosty
        tfidf_map[p.get("canonical")] = tfidf_keywords(soup.get_text(" ", strip=True))

        # simhash
        sim = simhash(tfidf_map[p.get("canonical")])
        near = [k for k,v in sim_index.items() if hamming(sim,v)<=4]  # bardzo podobne
        if near: warns.append("near-duplicate")
        sim_index[p.get("canonical")] = sim
        if "near-duplicate" in warns: dup_warns += 1

        # ZAPIS
        out_dir = OUT / lang / (slug or "")
        ensure_dir(out_dir)
        (out_dir/"index.html").write_text(str(soup), "utf-8")

        # do sitemap (tylko indexowalne)
        if not p.get("noindex"):
            indexables.append((p["canonical"], p.get("lastmod") or today, p.get("slugKey","home")))

        logs.append(f"{lang}/{slug or ''} [{p.get('__from','pages')}] warns={','.join(warns) if warns else '-'}")

    # Redirect stubs (z CMS.redirects)
    for r in CMS.get("redirects", []):
        src = r.get("from") or r.get("src") or ""
        dst = r.get("to")   or r.get("dst") or ""
        if not src or not dst: continue
        if not src.startswith("/"): src="/"+src
        if not src.endswith("/"): src+="/"
        dest = OUT / src.strip("/")
        ensure_dir(dest)
        write_text(dest/"index.html", f"<!doctype html><meta charset='utf-8'><meta http-equiv='refresh' content='0;url={dst}'><link rel='canonical' href='{dst}'><meta name='robots' content='noindex,follow'><title>Redirect</title>")

    # root redirect index (+ GSC meta + canonical)
    root_html = (
        f"<!doctype html><html lang=\"{DEFAULT_LANG}\"><head><meta charset=\"utf-8\">"
        f"<title>{CFG.get('site',{}).get('brand','Kras-Trans')}</title>"
        f"<meta name=\"google-site-verification\" content=\"{GSC}\">"
        f"<link rel=\"canonical\" href=\"/{DEFAULT_LANG}/\">"
        f"<meta http-equiv=\"refresh\" content=\"0; url=/{DEFAULT_LANG}/\">"
        f"<script>location.replace('/{DEFAULT_LANG}/');</script>"
        f"</head><body></body></html>"
    )
    write_text(OUT/"index.html", root_html)
    # GSC HTML file verification (drugi, pewny sposób weryfikacji)
    html_file = (CFG.get("constants", {}).get("GSC_HTML_FILE") or "").strip()
    if html_file and html_file.startswith("google") and html_file.endswith(".html"):
        # "google4377ff145fac0f52.html" -> token: "4377ff145fac0f52"
        token = html_file.replace("google", "", 1).removesuffix(".html")
        write_text(OUT / html_file, f"google-site-verification: {token}")

    # 404.html (prosty)
    write_text(OUT/"404.html", "<h1>404</h1><p>Nie znaleziono strony. <a href='/pl/'>Wróć do strony głównej</a>.</p>")

    # robots.txt
    write_text(OUT/"robots.txt", f"User-agent: *\nAllow: /\nSitemap: {SITE_URL}/sitemap.xml\n")

    # BingSiteAuth.xml
    if BING_USER:
        write_text(OUT/"BingSiteAuth.xml", f'<?xml version="1.0"?><users><user>{BING_USER}</user></users>')

    # IndexNow key file
    if INDEXNOW_KEY and INDEXNOW_KEY.lower()!="change_me_indexnow_key":
        write_text(OUT/f"{INDEXNOW_KEY}.txt", INDEXNOW_KEY)


    # SITEMAPY
    write_sitemaps(indexables, CMS.get("hreflang", {}))
    if NEWS_ENABLED or (CFG.get("blog",{}).get("news_sitemap",{}).get("enabled", False)):
        write_news_sitemap()

    # SEARCH INDEX (on-site)
    build_search_indexes()

    # FEEDS (RSS/Atom prosty)
    build_feeds()

    # Link-checker (wewnętrzny)
    internal_link_checker()

    # RAPORT
    report = [
        f"[OK] pages={len(pages)} city×service={len(city)} indexable={len(indexables)}",
        f"autolinks_inline={autolink_inline} fallback_cards={autolink_fb}",
        f"near_duplicates_warn={dup_warns}"
    ]
    write_text(OUT/"_reports"/"summary.txt", "\n".join(report))
    print("\n".join(report))
    print("\n".join(logs[:80] + (["…"] if len(logs)>80 else [])))

# ----------------------------- SITEMAPS ------------------------------------
def write_sitemaps(urls: List[Tuple[str, str, str]] | List[Tuple[str, str]] , alternates: Dict[str, Dict[str, str]] | None = None):
    # urls: [(loc, lastmod, slugKey?)] - trzeci element może nie wystąpić (wtedy alternates ignorujemy)
    # alternates: { slugKey: { 'pl': '...', 'en': '...', ... } }
    alternates = alternates or {}
    # wyciągnij slugKey jeśli jest; ujednolić do 3-elementowej krotki
    norm_urls: List[Tuple[str, str, str]] = []
    for u in urls:
        if len(u) == 3:
            norm_urls.append(u)  # (loc, lastmod, slugKey)
        elif len(u) == 2:
            loc, lastmod = u
            norm_urls.append((loc, lastmod, ""))  # brak slugKey
        else:
            continue

    shard = int(CFG.get("sitemap", {}).get("shard_size", 45000))
    groups = [norm_urls[i:i+shard] for i in range(0, len(norm_urls), shard)]
    index: List[Tuple[str, str]] = []

    for i, g in enumerate(groups or [[]]):
        name = f"sitemap-{i+1}.xml"
        path = OUT / name
        lines = ['<?xml version="1.0" encoding="UTF-8"?>',
                 '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" xmlns:xhtml="http://www.w3.org/1999/xhtml">']
        for loc, lastmod, slugKey in g:
            lines.append("  <url>")
            lines.append(f"    <loc>{loc}</loc>")
            lines.append(f"    <lastmod>{lastmod}</lastmod>")
            # alternates (hreflang) jeśli mamy slugKey i są wpisy
            if slugKey and CFG.get("sitemap", {}).get("include_alternates", True):
                for L, href in (alternates.get(slugKey, {}) or {}).items():
                    lines.append(f'    <xhtml:link rel="alternate" hreflang="{L}" href="{href}"/>')
            lines.append("  </url>")
        lines.append("</urlset>")
        write_text(path, "\n".join(lines))
        index.append((f"{SITE_URL}/{name}", UTC()))

    # index
    idx = ['<?xml version="1.0" encoding="UTF-8"?>',
           '<sitemapindex xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
    for loc, lastmod in index:
        idx.append(f"  <sitemap><loc>{loc}</loc><lastmod>{lastmod}</lastmod></sitemap>")
    idx.append("</sitemapindex>")
    write_text(OUT/"sitemap.xml", "\n".join(idx))

# ------------------------------ SEARCH INDEX -------------------------------
def build_search_indexes():
    docs_by_lang={L:[] for L in LOCALES}
    for lang_dir in (OUT).glob("*"):
        if not lang_dir.is_dir(): continue
        L=lang_dir.name
        if L not in LOCALES: continue
        for idx in lang_dir.rglob("index.html"):
            html=read_text(idx)
            s=soupify(html)
            h1=(s.find("h1").get_text(" ",strip=True) if s.find("h1") else "")
            title=(s.find("title").get_text(" ",strip=True) if s.find("title") else "")
            desc=""
            m=s.find("meta", attrs={"name":"description"})
            if m and m.get("content"): desc=m["content"]
            body=s.find("main") or s
            text=body.get_text(" ",strip=True)
            docs_by_lang[L].append({
                "path": "/"+idx.relative_to(OUT).as_posix().replace("index.html",""),
                "title": title, "h1": h1, "desc": desc, "text": text[:6000]
            })
    for L, arr in docs_by_lang.items():
        if not arr: continue
        write_text(OUT/f"search-index-{L}.json", json.dumps(arr, ensure_ascii=False))

# ------------------------------ FEEDS --------------------------------------
def build_feeds():
    posts=[p for p in CMS.get("pages",[]) if p.get("type")=="blog_post"]
    if not posts: return
    brand = CFG.get("site",{}).get("brand","Kras-Trans")
    for L in LOCALES:
        postsL=[p for p in posts if p.get("lang")==L]
        if not postsL: continue
        # RSS
        rss=['<?xml version="1.0" encoding="UTF-8"?>',
             '<rss version="2.0"><channel>',
             f"<title>{brand} – Blog</title>",
             f"<link>{SITE_URL}/{L}/blog/</link>",
             f"<description>Aktualności</description>"]
        for p in postsL[:50]:
            link=canonical(SITE_URL, L, p.get("slug",""), p.get("canonical_path"))
            rss.append("<item>")
            rss.append(f"<title>{(p.get('h1') or p.get('title') or '').strip()}</title>")
            rss.append(f"<link>{link}</link>")
            rss.append(f"<guid>{link}</guid>")
            rss.append(f"<description>{(p.get('meta_desc') or '').strip()}</description>")
            if p.get("date"):
                rss.append(f"<pubDate>{p.get('date')}</pubDate>")
            rss.append("</item>")
        rss.append("</channel></rss>")
        write_text(OUT/L/"feed.xml", "\n".join(rss))

        # Atom
        atom=['<?xml version="1.0" encoding="UTF-8"?>',
              '<feed xmlns="http://www.w3.org/2005/Atom">',
              f"<title>{brand} – Blog</title>",
              f"<link href=\"{SITE_URL}/{L}/blog/\"/>",
              f"<updated>{UTC()}</updated>",
              f"<id>{SITE_URL}/{L}/blog/</id>"]
        for p in postsL[:50]:
            link=canonical(SITE_URL, L, p.get("slug",""), p.get("canonical_path"))
            atom.append("<entry>")
            atom.append(f"<title>{(p.get('h1') or p.get('title') or '').strip()}</title>")
            atom.append(f"<link href=\"{link}\"/>")
            atom.append(f"<id>{link}</id>")
            atom.append(f"<updated>{p.get('date') or UTC()}</updated>")
            atom.append(f"<summary>{(p.get('meta_desc') or '').strip()}</summary>")
            atom.append("</entry>")
        atom.append("</feed>")
        write_text(OUT/L/"atom.xml", "\n".join(atom))

# ------------------------------ LINK-CHECKER -------------------------------
def internal_link_checker():
    all_paths=set()
    for idx in OUT.rglob("index.html"):
        url="/"+idx.relative_to(OUT).as_posix().replace("index.html","")
        all_paths.add(url)
    broken=[]
    for idx in OUT.rglob("index.html"):
        html=read_text(idx)
        s=soupify(html)
        for a in s.find_all("a", href=True):
            href=a["href"]
            if href.startswith("mailto:") or href.startswith("tel:"): continue
            if href.startswith("http"): continue
            if not href.endswith("/"): href=href+"/"
            if href not in all_paths:
                broken.append((str(idx.relative_to(OUT)), href))
    if broken:
        lines=["BROKEN INTERNAL LINKS (first 100):"] + [f"{p} → {h}" for p,h in broken[:100]]
        write_text(OUT/"_reports"/"broken-links.txt", "\n".join(lines))

# ----------------------------- NEWS SITEMAP ---------------------------------
def write_news_sitemap():
    # Generuje news-sitemap na podstawie CMS.pages[type=blog_post] w oknie czasowym.
    # Działa tylko gdy NEWS_ENABLED=True lub blog.news_sitemap.enabled=True.
    enabled_cfg = bool(CFG.get("blog", {}).get("news_sitemap", {}).get("enabled", False))
    if not (NEWS_ENABLED or enabled_cfg):
        return  # nic do roboty

    window_h = int(CFG.get("blog", {}).get("news_sitemap", {}).get("window_hours", 48))
    limit_dt = datetime.now(timezone.utc) - timedelta(hours=window_h)

    items = []
    for p in CMS.get("pages", []):
        if (p.get("type") == "blog_post") and p.get("date"):
            try:
                dt = datetime.fromisoformat(str(p["date"]).replace("Z", "+00:00"))
                if dt >= limit_dt:
                    loc = canonical(SITE_URL, p.get("lang", "pl"), p.get("slug", ""), p.get("canonical_path"))
                    items.append((loc, dt.isoformat()))
            except Exception:
                pass

    if not items:
        return

    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" xmlns:news="http://www.google.com/schemas/sitemap-news/0.9">'
    ]
    for loc, dt in items:
        lines.append("  <url>")
        lines.append(f"    <loc>{loc}</loc>")
        lines.append('    <news:news>')
        lines.append('      <news:publication><news:name>Kras-Trans</news:name><news:language>pl</news:language></news:publication>')
        lines.append(f"      <news:publication_date>{dt}</news:publication_date>")
        lines.append("      <news:title>Aktualność</news:title>")
        lines.append("    </news:news>")
        lines.append("  </url>")
    lines.append("</urlset>")

    write_text(OUT / "news-sitemap.xml", "\n".join(lines))


# ------------------------------ MAIN ---------------------------------------
if __name__=="__main__":
    build_all()
