#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate static NAV bundles from single XLSX:
  input:  data/cms.xlsx  (sheets: Routes, Nav, [Props])
  output: assets/nav/bundle_<lang>.json

- Zero GAS, zero limitów. Szybkie i powtarzalne.
- Primary bar = rodzice (bez href) + single (bez parent)
- Mega = dzieci pogrupowane po 'parent'
- Parent bez landing page => href="#", pełni rolę toggle (bez 404)
- CTA/logo/status/social z opcjonalnego sheeta Props (per lang)
"""

import json, sys, pathlib, re
from collections import defaultdict

from openpyxl import load_workbook
from unidecode import unidecode

LOCALES = ['pl','en','de','fr','it','ru','ua']

def slugify(raw: str) -> str:
    s = unidecode((raw or '').strip().lower())
    s = re.sub(r'[^a-z0-9]+', '-', s)
    s = re.sub(r'-{2,}', '-', s).strip('-')
    return s

def _to_bool(v: str) -> bool:
    s = str(v or '').strip().lower()
    return s in {'true','1','yes','y','tak','prawda'}

def read_routes(ws):
    start_cell, end_cell = ws.calculate_dimension().split(":")
    max_row = int("".join(filter(str.isdigit, end_cell)))
    head = [c.value for c in ws[1]]
    try: idx_slug = head.index('slugKey')
    except ValueError:
        raise SystemExit("[Routes] Missing 'slugKey' header")
    langs = [h for h in head if h and h.lower() in LOCALES]
    routes = {}
    for row in ws.iter_rows(min_row=2, values_only=True, max_row=max_row):
        slug_key = (row[idx_slug] or '').strip()
        if not slug_key: continue
        m = {}
        for h, v in zip(head, row):
            if not h: continue
            k = h.lower()
            if k in LOCALES:
                m[k] = (v or '').strip().lstrip('/').rstrip('/')
        routes[slug_key] = m
    return routes

def read_props(ws):
    start_cell, end_cell = ws.calculate_dimension().split(":")
    max_row = int("".join(filter(str.isdigit, end_cell)))
    props = defaultdict(dict)  # props[lang][key] = value
    head = [c.value for c in ws[1]]
    try:
        i_key  = head.index('key')
        i_lang = head.index('lang')
        i_val  = head.index('value')
    except ValueError:
        raise SystemExit("[Props] Expected headers: key, lang, value")
    for row in ws.iter_rows(min_row=2, values_only=True, max_row=max_row):
        key  = (row[i_key]  or '').strip()
        lang = (row[i_lang] or '').strip().lower()
        val  = (row[i_val]  or '').strip()
        if key and lang:
            props[lang][key] = val
    return props

def read_nav(ws):
    start_cell, end_cell = ws.calculate_dimension().split(":")
    max_row = int("".join(filter(str.isdigit, end_cell)))
    head = [c.value for c in ws[1]]
    idx = {}
    for h in ('lang','label','href','parent','order','enabled'):
        try:
            idx[h] = head.index(h)
        except ValueError:
            raise SystemExit(f"[Nav] Missing header '{h}'")
    if 'col' in head:
        idx['col'] = head.index('col')

    per_lang = defaultdict(list)
    seen_href = defaultdict(set)
    for row in ws.iter_rows(min_row=2, values_only=True, max_row=max_row):
        lang = (row[idx['lang']] or '').strip().lower()
        if lang not in LOCALES:
            continue

        label  = (row[idx['label']] or '').strip()
        href   = (row[idx['href']]  or '').strip()
        parent = (row[idx['parent']] or '').strip()
        order  = row[idx['order']]
        enabled= row[idx['enabled']]
        col    = None
        if 'col' in idx:
            try:
                c = int(row[idx['col']])
                if 1 <= c <= 4:
                    col = c
            except Exception:
                pass

        if not label:
            continue
        if not _to_bool(enabled):
            continue
        if href:
            if href in seen_href[lang]:
                print(f"[Nav] Duplicate href '{href}' in {lang}", file=sys.stderr)
                continue
            seen_href[lang].add(href)
        try:
            order = int(order)
        except Exception:
            order = 0

        per_lang[lang].append({
            'label': label,
            'href': href,
            'parent': parent,
            'order': order,
            'col': col,
        })

    return per_lang

def langs_html_for_lang(lang):
    # prosty picker języków – link na /{L}/ + aria-current
    tags = []
    for L in LOCALES:
        cur = ' aria-current="true"' if L==lang else ''
        tags.append(f'<a href="/{L}/"{cur}><img class="flag" src="/assets/flags/{ "gb" if L=="en" else L }.svg" alt="" width="18" height="12">{L.upper()}</a>')
    return ' '.join(tags)

def build_bundle(lang, rows, routes, props):
    # Structure
    tops = [r for r in rows if not (r.get('parent') or '').strip()]
    children = defaultdict(list)
    for r in rows:
        p = (r.get('parent') or '').strip()
        if p:
            children[p].append(r)

    for arr in [tops, *children.values()]:
        arr.sort(key=lambda r: int(r.get('order') or 0))

    # Primary HTML
    li = []
    for t in tops:
        label = t['label']
        href  = t['href']
        slug  = slugify(label)
        kids  = children.get(label, []) + children.get(slug, [])
        dp    = f' data-panel="{label}"' if kids else ''
        li.append(f'<li{dp}><a href="{href}">{label}</a></li>')
    primary_html = ''.join(li)

    # Mega HTML
    sections = []
    for t in tops:
        label = t['label']
        slug  = slugify(label)
        kids  = children.get(label, []) + children.get(slug, [])
        if not kids:
            continue
        cols = {i: [] for i in range(1, 5)}
        for ch in kids:
            try:
                c = int(ch.get('col') or 1)
            except Exception:
                c = 1
            if c not in cols:
                c = 1
            cols[c].append(ch)
        col_divs = []
        for i in range(1, 5):
            links = ''.join(f'<div><a href="{r["href"]}">{r["label"]}</a></div>' for r in cols[i])
            col_divs.append(f'<div class="col">{links}</div>')
        sections.append(
            f'<section class="mega__section" data-panel="{label}"><div class="mega__grid">' + ''.join(col_divs) + '</div></section>'
        )
    mega_html = ''.join(sections)

    bundle = {
        'primary_html': primary_html,
        'mega_html':    mega_html,
        'langs_html':   langs_html_for_lang(lang),
        'routes':       routes,
    }
    # Props (opcjonalnie)
    p = props.get(lang, {})
    if p:
        # CTA
        cta = {}
        if p.get('cta_label'):   cta['label']   = p['cta_label']
        if p.get('cta_slugKey'): cta['slugKey'] = p['cta_slugKey']
        if cta: bundle['cta'] = cta
        # Logo
        logo = {}
        if p.get('logo_src'): logo['src'] = p['logo_src']
        if p.get('logo_alt'): logo['alt'] = p['logo_alt']
        if logo: bundle['logo'] = logo
        # Status
        status = {}
        if p.get('status_label'): status['label'] = p['status_label']
        if p.get('status_href'):  status['href']  = p['status_href']
        if status: bundle['status'] = status
        # Social
        social = {}
        for k in ('ig','li','fb'):
            v = p.get(f'social_{k}')
            if v: social[k] = v
        if social: bundle['social'] = social
    return bundle
  
def main():
    src = pathlib.Path('data/cms.xlsx')
    if not src.exists(): raise SystemExit("Missing data/cms.xlsx")
    wb = load_workbook(str(src), data_only=True)

    # Required sheets
    if 'Routes' not in wb.sheetnames or 'Nav' not in wb.sheetnames:
        raise SystemExit("Expected sheets: 'Routes' and 'Nav' (optional: 'Props')")

    routes = read_routes(wb['Routes'])
    nav_map = read_nav(wb['Nav'])
    props   = read_props(wb['Props']) if 'Props' in wb.sheetnames else {}

    out_dir = pathlib.Path('assets/nav')
    out_dir.mkdir(parents=True, exist_ok=True)

    total = 0
    for lang in LOCALES:
        items = nav_map.get(lang, [])
        if not items: continue
        bundle = build_bundle(lang, items, routes, props)
        out = out_dir / f"bundle_{lang}.json"
        out.write_text(json.dumps(bundle, ensure_ascii=False, separators=(',',':')), encoding='utf-8')
        print(f"✔ {out} ({out.stat().st_size} bytes)")
        total += 1
    print(f"Done. Generated {total} bundles in assets/nav/")
    return 0

if __name__ == "__main__":
    sys.exit(main())
